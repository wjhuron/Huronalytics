"""audit_write.py — render pitch_tag_audit results to an .xlsx workbook."""
import os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from pitch_tag_audit import deg_clock

HDR_FILL = PatternFill('solid', fgColor='1F3864')
HDR_FONT = Font(bold=True, color='FFFFFF')
HIGH_FILL = PatternFill('solid', fgColor='C6E0B4')   # green
MED_FILL = PatternFill('solid', fgColor='FFE699')    # amber


def _r1(v):
    return round(v, 1) if isinstance(v, (int, float)) else v


def _style_header(ws, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal='center')
    ws.freeze_panes = 'A2'


def _autosize(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


G1_HDR = ['Pitcher', 'Date', 'Opp', 'Current', 'Suggested', 'Conf', 'Tier',
          'GameFlips', 'Velo', 'IVB', 'HB', 'Spin', 'RTilt', 'OTilt', 'Arm',
          'Why', 'PitchID']
G1_W = [22, 11, 5, 8, 10, 6, 8, 10, 7, 7, 7, 7, 7, 7, 7, 60, 16]

G2_HDR = ['Team', 'Pitcher', 'Hand', 'Type', 'N', 'Suggested', 'Conf', 'Tier',
          'Velo', 'IVB', 'HB', 'Spin', 'RTilt', 'OTilt', 'Arm', 'Why']
G2_W = [6, 22, 5, 6, 6, 10, 6, 8, 7, 7, 7, 7, 7, 7, 7, 62]


def _g1_row(f):
    mv = f['mv']
    return [
        f['pitcher'], f['date'], f['opp'], f['own'], f['tgt'], f['conf'],
        f['tier'], f"{f['nflip']}/{f['ntype']}",
        _r1(mv.get('Velocity')), _r1(mv.get('IndVertBrk')), _r1(mv.get('HorzBrk')),
        round(mv['Spin Rate']) if mv.get('Spin Rate') is not None else '',
        deg_clock(mv.get('RTilt')), deg_clock(mv.get('OTilt')),
        _r1(mv.get('ArmAngle')), f['why'], f['pid'],
    ]


def _g2_row(x):
    c = x['cen']
    return [
        x['team'], x['pitcher'], x['hand'], x['type'], x['n'], x['tgt'],
        x['conf'], x['tier'],
        _r1(c.get('Velocity')), _r1(c.get('IndVertBrk')), _r1(c.get('HorzBrk')),
        round(c['Spin Rate']) if c.get('Spin Rate') is not None else '',
        deg_clock(c.get('RTilt')), deg_clock(c.get('OTilt')),
        _r1(c.get('ArmAngle')), x['why'],
    ]


def _tier_fill(ws, row, tier, ncols):
    fill = HIGH_FILL if tier == 'High' else MED_FILL if tier == 'Medium' else None
    if fill:
        ws.cell(row=row, column=7 if ncols == len(G1_HDR) else 8).fill = fill


def write_workbook(path, g1, g2):
    wb = Workbook()

    # ---- README ----
    ws = wb.active
    ws.title = 'README'
    lines = [
        ('Pitch Tag Audit 2026', True),
        ('', False),
        ('Two audits of the retagged Google Sheets pitch classifications.', False),
        ('', False),
        ('GOAL 1 tabs (one per team): individual pitches likely mistagged.', True),
        ('For each pitcher, stable season per-type centroids are built, then every', False),
        ('pitch is tested: is it closer to a DIFFERENT type this pitcher throws than', False),
        ('to its own tag? Metrics: velocity, spin, RTilt, OTilt (spin-direction,', False),
        ('circular), IVB, HB, arm angle; each standardized by its natural within-', False),
        ('pitch noise. "GameFlips" = how many pitches of that tag in that game flipped', False),
        ('the same way / total of that tag thrown that game. A high fraction is a real', False),
        ('within-game trend (a whole game or stretch mislabeled), not tracking noise.', False),
        ('', False),
        ('GOAL 2 tab (Whole-Type Reclass): entire pitch types to reclassify.', True),
        ('Each pitcher x type cluster is compared to league per-hand type prototypes;', False),
        ('flagged when the cluster matches a different label better (e.g. Rico Garcia', False),
        ('SL -> FC). FF<->SI swaps excluded. A ->FC suggestion also requires the pitch', False),
        ('to be within 7.5 mph of the pitcher\'s own fastball (a cutter sits close to', False),
        ('the FB; a gyro slider is much slower even when its shape looks cutter-like).', False),
        ('A type is never suggested to merge into one the pitcher already throws as its', False),
        ('own pitch (a distinct existing FC means the SL is a separate pitch, not the FC).', False),
        ('', False),
        ('Confidence: High = strong, minimal review; Medium = worth a look. Low is', True),
        ('filtered out. Goal 1 confidence blends margin, how many metrics agree, and', False),
        ('within-game reinforcement. Goal 2 confidence is driven by the profile gap.', False),
        ('', False),
        ('Tilt shown as clock (H:MM). "Why" names the metrics that most drive the', False),
        ('suggestion, with (currentTag value / suggestedTag value) for context.', False),
        ('READ-ONLY: nothing here is applied to the sheets. Wally makes the final call.', True),
    ]
    for i, (txt, bold) in enumerate(lines, 1):
        cell = ws.cell(row=i, column=1, value=txt)
        if bold:
            cell.font = Font(bold=True, size=12 if i == 1 else 11)
    ws.column_dimensions['A'].width = 95

    # ---- Goal 2 (single tab) ----
    ws2 = wb.create_sheet('Whole-Type Reclass')
    ws2.append(G2_HDR)
    for x in sorted(g2, key=lambda x: (x['tier'] != 'High', x['team'] or '',
                                       x['pitcher'])):
        if x['tier'] == 'Low':
            continue
        ws2.append(_g2_row(x))
        _tier_fill(ws2, ws2.max_row, x['tier'], len(G2_HDR))
    _style_header(ws2, len(G2_HDR))
    _autosize(ws2, G2_W)

    # ---- Goal 1 (one tab per team) ----
    byteam = defaultdict(list)
    for f in g1:
        if f['tier'] == 'Low':
            continue
        byteam[f['team'] or 'UNK'].append(f)
    for team in sorted(byteam):
        ws = wb.create_sheet(team[:31])
        ws.append(G1_HDR)
        rows = sorted(byteam[team],
                      key=lambda f: (f['pitcher'], f['date'] or '', -f['conf']))
        for f in rows:
            ws.append(_g1_row(f))
            _tier_fill(ws, ws.max_row, f['tier'], len(G1_HDR))
        _style_header(ws, len(G1_HDR))
        _autosize(ws, G1_W)

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    nteams = len(byteam)
    ng1 = sum(len(v) for v in byteam.values())
    ng2 = sum(1 for x in g2 if x['tier'] != 'Low')
    print(f"  Goal 1: {ng1} pitch flags across {nteams} team tabs")
    print(f"  Goal 2: {ng2} whole-type flags")
