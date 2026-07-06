"""build_tracking_mismatch_xlsx.py — workbook of tracking-field discrepancies vs
current Savant. READ-ONLY. One "clearly wrong" tab + one "(all)" tab per metric.

Mappings + signs verified empirically (median diff ~0 on aligned pitches):
  Velocity=release_speed, Spin Rate=release_spin_rate, Extension=release_extension,
  IndVertBrk=pfx_z*12, HorzBrk=-pfx_x*12, PlateX=plate_x, PlateZ=plate_z,
  SzTop=sz_top, SzBot=sz_bot, RelPosX=release_pos_x, RelPosZ=release_pos_z,
  ExitVelo=launch_speed, LaunchAngle=launch_angle, Distance=hit_distance_sc,
  HC_X=hc_x, HC_Y=hc_y.
Each field's systematic feed-vs-Savant median offset is removed before flagging.
Fields are categorised:
  FIXABLE  — feed≈Savant for ~all pitches, small genuine error tail.
  SCATTER  — feed and Savant differ per-pitch by design (release point measured on
             the 50-ft plane vs extension-adjusted); differences are NOT errors.
Pitches matched with auto-ball-aware numbering. Output:
  ~/Downloads/tracking_mismatches_2026.xlsx

Usage: python3 scripts/build_tracking_mismatch_xlsx.py            (uses cache if present)
       python3 scripts/build_tracking_mismatch_xlsx.py --reread   (force re-read sheets)
"""
import os, sys, time, pickle, statistics
from collections import defaultdict
ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)
import gspread
import backfill_supplement as B
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

CACHE = os.path.join(ROOT, 'data', '_tracking_pitches.pkl')
CAP = 20000  # max rows per sheet (Excel-friendly; scatter fields can exceed this)

# col, savant_attr, xf(sv), clear_thr, all_floor, kind
FIELDS = [
    ('Spin Rate',   'release_spin_rate', lambda v: v,       75,   2,    'FIXABLE'),
    ('Velocity',    'release_speed',     lambda v: v,       0.7,  0.15, 'FIXABLE'),
    ('IndVertBrk',  'pfx_z',             lambda v: v * 12,  1.5,  0.15, 'FIXABLE'),
    ('HorzBrk',     'pfx_x',             lambda v: -v * 12, 1.5,  0.15, 'FIXABLE'),
    ('Extension',   'release_extension', lambda v: v,       0.25, 0.05, 'FIXABLE'),
    ('PlateX',      'plate_x',           lambda v: v,       0.20, 0.04, 'FIXABLE'),
    ('SzTop',       'sz_top',            lambda v: v,       0.15, 0.04, 'FIXABLE'),
    ('SzBot',       'sz_bot',            lambda v: v,       0.15, 0.04, 'FIXABLE'),
    ('ExitVelo',    'launch_speed',      lambda v: v,       2.0,  0.2,  'FIXABLE'),
    ('LaunchAngle', 'launch_angle',      lambda v: v,       3.0,  1.0,  'FIXABLE'),
    ('Distance',    'hit_distance_sc',   lambda v: v,       12,   2,    'FIXABLE'),
    ('HC_X',        'hc_x',              lambda v: v,       3.0,  1.0,  'FIXABLE'),
    ('HC_Y',        'hc_y',              lambda v: v,       3.0,  1.0,  'FIXABLE'),
    ('RelPosX',     'release_pos_x',     lambda v: v,       0.6,  0.10, 'SCATTER'),
    ('RelPosZ',     'release_pos_z',     lambda v: v,       0.5,  0.10, 'SCATTER'),
    ('PlateZ',      'plate_z',           lambda v: v,       0.35, 0.05, 'SCATTER'),
]


def sf(x):
    try:
        v = float(x)
        return v if v == v else None
    except (TypeError, ValueError):
        return None


def build_aligned():
    sc = pickle.load(open(os.path.join(ROOT, 'data', '_statcast2026_full.pkl'), 'rb'))
    bypa = defaultdict(list)
    for r in sc.itertuples(index=False):
        try:
            bypa[(int(r.game_pk), int(r.at_bat_number))].append((int(r.pitch_number), r))
        except Exception:
            continue
    aligned = {}
    for (pk, ab), evs in bypa.items():
        evs.sort(key=lambda t: t[0])
        feed = 0
        for pn, r in evs:
            if 'automatic' in str(r.description or '').lower():
                continue
            feed += 1
            aligned[(pk, ab, feed)] = {a: sf(getattr(r, a, None)) for _, a, *_ in FIELDS}
    return aligned


def read_pitches(aligned):
    gc = gspread.service_account()
    out = []
    for label, sid in B.SPREADSHEET_IDS.items():
        sh = gc.open_by_key(sid)
        for ws in sh.worksheets():
            t = ws.title.upper()
            if t not in B.ALL_TRACKED_TEAMS or t in ('ROC', 'AAA', 'FCL'):
                continue
            time.sleep(0.7)
            vals = ws.get_all_values(value_render_option='UNFORMATTED_VALUE')
            if not vals or len(vals) < 2 or 'PitchID' not in vals[0]:
                continue
            ci = {n: j for j, n in enumerate(vals[0]) if n}
            for r in vals[1:]:
                pid = str(r[ci['PitchID']]) if ci['PitchID'] < len(r) else ''
                pa = pid.split('_')
                if len(pa) != 3:
                    continue
                try:
                    key = (int(pa[0]), int(pa[1]), int(pa[2]))
                except ValueError:
                    continue
                sav = aligned.get(key)
                if sav is None:
                    continue
                rec = {'pid': pid, 'sav': sav,
                       'pitcher': r[ci['Pitcher']] if 'Pitcher' in ci else '',
                       'ptype': r[ci['Pitch Type']] if 'Pitch Type' in ci else '',
                       'date': r[ci['Game Date']] if 'Game Date' in ci else ''}
                for col, *_ in FIELDS:
                    rec[col] = sf(r[ci[col]]) if col in ci and ci[col] < len(r) else None
                out.append(rec)
            print(f"  [{label}/{ws.title}] read", flush=True)
    return out


def main():
    aligned = build_aligned()
    print(f"aligned savant pitches: {len(aligned)}", flush=True)
    if os.path.exists(CACHE) and '--reread' not in sys.argv:
        pitches = pickle.load(open(CACHE, 'rb'))
        print(f"loaded {len(pitches)} pitches from cache", flush=True)
    else:
        pitches = read_pitches(aligned)
        pickle.dump(pitches, open(CACHE, 'wb'))
        print(f"read + cached {len(pitches)} pitches", flush=True)

    # pitcher×pitchtype averages per field for plausibility
    favg = {col: defaultdict(lambda: [0.0, 0]) for col, *_ in FIELDS}
    for rec in pitches:
        for col, *_ in FIELDS:
            v = rec[col]
            if v is not None:
                a = favg[col][(rec['pitcher'], rec['ptype'])]
                a[0] += v; a[1] += 1
    def tavg(col, pit, pt):
        a = favg[col][(pit, pt)]
        return round(a[0] / a[1], 1) if a[1] >= 5 else ''

    # fields where the pitcher's pitch-type average is a valid "expected" reference,
    # so we can say which of {sheet, savant} is the outlier.
    REF_FIELDS = {'Spin Rate', 'Velocity', 'IndVertBrk', 'HorzBrk', 'Extension'}

    def verdict(col, sheet, savant, avg):
        if col not in REF_FIELDS or avg in ('', None):
            return ''
        ds, dv = abs(sheet - avg), abs(savant - avg)
        if ds < dv * 0.6:
            return 'KEEP SHEET (Savant regressed)'
        if dv < ds * 0.6:
            return 'USE SAVANT'
        return 'REVIEW'

    wb = Workbook(); wb.remove(wb.active)
    hdr = ['PitchID', 'Date', 'Pitcher', 'PitchType', 'Sheet', 'Savant', 'Diff', 'PitcherTypeAvg', 'Verdict']
    yellow = PatternFill('solid', fgColor='FFF2CC')
    summary = []

    def add_sheet(name, rows, note=''):
        ws = wb.create_sheet(title=name[:31])
        if note:
            ws.append([note]); ws.append([])
        ws.append(hdr)
        hr = ws.max_row
        for c in range(1, len(hdr) + 1):
            ws.cell(hr, c).font = Font(bold=True)
        for row in rows[:CAP]:
            ws.append(list(row))
        return ws

    for col, attr, xf, clear_thr, floor, kind in FIELDS:
        offs = []
        for rec in pitches:
            pv = rec[col]; sv = rec['sav'].get(attr)
            if pv is not None and sv is not None:
                offs.append(pv - xf(sv))
        med_off = statistics.median(offs) if offs else 0.0
        clear_rows, all_rows = [], []
        for rec in pitches:
            pv = rec[col]; sv = rec['sav'].get(attr)
            if pv is None or sv is None:
                continue
            target = xf(sv)
            resid = (pv - target) - med_off
            if abs(resid) <= floor:
                continue
            avg = tavg(col, rec['pitcher'], rec['ptype'])
            r = (rec['pid'], rec['date'], rec['pitcher'], rec['ptype'],
                 round(pv, 2), round(target, 2), round(pv - target, 2),
                 avg, verdict(col, pv, target, avg))
            all_rows.append((abs(resid), r))
            if abs(resid) > clear_thr:
                clear_rows.append((abs(resid), r))
        clear_rows.sort(key=lambda x: -x[0]); all_rows.sort(key=lambda x: -x[0])
        clear_rows = [r for _, r in clear_rows]; all_rows = [r for _, r in all_rows]
        summary.append((col, kind, len(clear_rows), len(all_rows), round(med_off, 3)))
        cav = ('' if kind == 'FIXABLE' else
               'SCATTER: feed vs Savant differ per-pitch by measurement design (50-ft plane vs '
               'extension-adjusted); these are NOT errors — review before changing.')
        if clear_rows:
            add_sheet(f"{col} wrong", clear_rows, cav)
        if all_rows:
            capnote = f' (showing worst {CAP} of {len(all_rows)})' if len(all_rows) > CAP else ''
            add_sheet(f"{col} all", all_rows, (cav + capnote) if (cav or capnote) else '')

    sm = wb.create_sheet(title='SUMMARY', index=0)
    sm.append(['Metric', 'Category', 'ClearlyWrong', 'AllDiscrepancies', 'SysOffset(removed)'])
    for c in range(1, 6):
        sm.cell(1, c).font = Font(bold=True)
    for col, kind, nc, na, off in summary:
        sm.append([col, kind, nc, na, off])
        if kind == 'SCATTER':
            for c in range(1, 6):
                sm.cell(sm.max_row, c).fill = yellow
    sm.append([])
    sm.append(['matched pitches', len(pitches)])
    sm.append(['FIXABLE', 'feed matches Savant for ~all pitches; a diff = genuine stale/wrong value. "Savant" col = the value to write.'])
    sm.append(['SCATTER (yellow)', 'release point / plate-Z differ per pitch by measurement convention, not error. Do not bulk-change.'])

    out = os.path.expanduser('~/Downloads/tracking_mismatches_2026.xlsx')
    wb.save(out)
    print(f"\nwrote {out}\n")
    print(f"{'metric':12s} {'kind':8s} {'wrong':>6s} {'all':>7s} {'offset':>8s}")
    for col, kind, nc, na, off in summary:
        print(f"{col:12s} {kind:8s} {nc:6d} {na:7d} {off:8.3f}")


if __name__ == '__main__':
    main()
