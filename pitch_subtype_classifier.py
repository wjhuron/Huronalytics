#!/usr/bin/env python3
"""
Pitch Subtype Classifier
Classifies every pitcher's FF/SI/FC/SL/CU/ST/SV pitches into specific subtypes
based on IVB (Induced Vertical Break) and HB (Horizontal Break) using
nearest-neighbor matching to reference movement profiles.

For LHP: HB is flipped to normalize to RHP perspective before classification.

Reference profiles from movement charts (RHP perspective):
  - Positive HB = arm-side run
  - Negative HB = glove-side break
"""

import json, math
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Reference subtype profiles (IVB, HB) — RHP perspective ───────────────
# Each entry: (subtype_name, base_type, ivb, hb)

SUBTYPES = [
    # ── FF (Four-Seam Fastball) ──
    ('Gyro Fastball',    'FF',   9,     0),
    ('Inefficient FF',   'FF',  13,     6),
    ('Deadzone FB',      'FF',  14,    13.5),
    ('Running Fastball',  'FF',  15,    16),
    ('Relative Cut FF',  'FF',  16,     1.5),
    ('Standard FF',      'FF',  17,     9),
    ('Rider',            'FF',  19.5,   6.5),
    ('Ride n\' Run',     'FF',  19.5,  11),

    # ── SI (Sinker) ──
    ('Gyro Sinker',      'SI',   9,    12),
    ('Sinker',           'SI',   9,    16),
    ('Running Sinker',   'SI',   9.5,  19.5),
    ('Heavy Sinker',     'SI',   5,    15.5),
    ('Heavy Runner',     'SI',   5,    20),
    ('Diver',            'SI',  -1,    17),

    # ── FC (Cutter) ──
    ('Gyro Cutter',      'FC',   8,     0),
    ('Standard Cutter',  'FC',  10,    -3),
    ('Sweeping Cutter',  'FC',   9.5,  -6.5),
    ('Backspinner',      'FC',  13.5,   0.5),

    # ── SL (Slider) ──
    ('Gyro SL',          'SL',   1,    -2),
    ('Slutter',          'SL',   5,    -4.5),
    ('Standard SL',      'SL',   0.5,  -6.5),

    # ── ST (Sweeper) ──
    ('Sweeper',          'ST',  -1.5, -15),

    # ── CU (Curveball) ──
    ('Gyro CB',          'CU',  -5,    -2),
    ('IE Downer',        'CU',  -9,    -2),
    ('Standard CB',      'CU', -13,   -10),
    ('Downer',           'CU', -15,    -5),
    ('Efficient CB',     'CU', -18,   -14),

    # ── SV (Slurve) ──
    ('IE Slurve',        'SV',  -6.5,  -7),
    ('Slurve',           'SV',  -8,   -14),
    ('Efficient Slurve', 'SV', -12,   -19),
]

TARGET_TYPES = {'FF', 'SI', 'FC', 'SL', 'ST', 'CU', 'SV'}


def euclidean_dist(ivb1, hb1, ivb2, hb2):
    return math.sqrt((ivb1 - ivb2) ** 2 + (hb1 - hb2) ** 2)


def classify_pitch(ivb, hb, hand):
    """
    Classify a pitch into the nearest subtype.
    For LHP, flip HB to normalize to RHP perspective.
    Returns (subtype_name, base_type, distance).
    """
    # Normalize to RHP perspective
    norm_hb = hb if hand == 'R' else -hb

    best_name = None
    best_base = None
    best_dist = float('inf')

    for name, base, ref_ivb, ref_hb in SUBTYPES:
        d = euclidean_dist(ivb, norm_hb, ref_ivb, ref_hb)
        if d < best_dist:
            best_dist = d
            best_name = name
            best_base = base

    return best_name, best_base, round(best_dist, 2)


def get_second_closest(ivb, hb, hand, exclude_name):
    """Get second-closest subtype for context."""
    norm_hb = hb if hand == 'R' else -hb
    best_name = None
    best_base = None
    best_dist = float('inf')

    for name, base, ref_ivb, ref_hb in SUBTYPES:
        if name == exclude_name:
            continue
        d = euclidean_dist(ivb, norm_hb, ref_ivb, ref_hb)
        if d < best_dist:
            best_dist = d
            best_name = name
            best_base = base

    return best_name, best_base, round(best_dist, 2)


# ── Load data ─────────────────────────────────────────────────────────────
with open('data/pitch_leaderboard.json') as f:
    pitch_data = json.load(f)

print(f"Loaded {len(pitch_data)} pitch rows")

# ── Classify every relevant pitch ────────────────────────────────────────
results = []

for p in pitch_data:
    pt = p.get('pitchType', '')
    if pt not in TARGET_TYPES:
        continue

    ivb = p.get('indVertBrk')
    hb = p.get('horzBrk')
    if ivb is None or hb is None:
        continue

    hand = p.get('throws', 'R')
    pitcher = p.get('pitcher') or p.get('name', '?')
    team = p.get('team', '?')
    velo = p.get('velocity')
    spin = p.get('spinRate')
    count = p.get('count', 0)
    tilt = p.get('breakTilt', '')

    subtype, suggested_base, dist = classify_pitch(ivb, hb, hand)
    second_name, second_base, second_dist = get_second_closest(ivb, hb, hand, subtype)

    changed = (suggested_base != pt)

    results.append({
        'pitcher': pitcher,
        'team': team,
        'hand': hand,
        'current_type': pt,
        'subtype': subtype,
        'suggested_base': suggested_base,
        'changed': changed,
        'distance': dist,
        'second_subtype': second_name,
        'second_base': second_base,
        'second_distance': second_dist,
        'ivb': ivb,
        'hb': hb,
        'velocity': velo,
        'spin': spin,
        'count': count,
        'tilt': tilt,
    })

# Sort by pitcher name, then current type
results.sort(key=lambda r: (r['pitcher'], r['current_type']))

total = len(results)
changed_count = sum(1 for r in results if r['changed'])
print(f"\nClassified {total} pitches")
print(f"  Base type matches current label: {total - changed_count}")
print(f"  Base type DIFFERS from current label: {changed_count}")

# ── Subtype distribution ──
from collections import Counter
subtype_counts = Counter(r['subtype'] for r in results)
print("\nSubtype distribution:")
for name, cnt in subtype_counts.most_common():
    base = next(b for n, b, _, _ in SUBTYPES if n == name)
    print(f"  {name} ({base}): {cnt}")

# Change breakdown
if changed_count:
    change_counts = Counter(f"{r['current_type']} → {r['suggested_base']}" for r in results if r['changed'])
    print(f"\nReclassification breakdown:")
    for change, cnt in change_counts.most_common():
        print(f"  {change}: {cnt}")


# ══════════════════════════════════════════════════════════════════════════
#  Excel Output
# ══════════════════════════════════════════════════════════════════════════
wb = Workbook()

# ── Styles ──
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='1a1a2e', end_color='1a1a2e', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

data_font = Font(name='Calibri', size=10)
bold_font = Font(name='Calibri', size=10, bold=True)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

# Color fills by base type
BASE_COLORS = {
    'FF': PatternFill(start_color='DBEAFE', end_color='DBEAFE', fill_type='solid'),  # Light blue
    'SI': PatternFill(start_color='FDE68A', end_color='FDE68A', fill_type='solid'),  # Yellow
    'FC': PatternFill(start_color='C4B5FD', end_color='C4B5FD', fill_type='solid'),  # Purple
    'SL': PatternFill(start_color='A7F3D0', end_color='A7F3D0', fill_type='solid'),  # Green
    'ST': PatternFill(start_color='FBCFE8', end_color='FBCFE8', fill_type='solid'),  # Pink
    'CU': PatternFill(start_color='FED7AA', end_color='FED7AA', fill_type='solid'),  # Orange
    'SV': PatternFill(start_color='E5E7EB', end_color='E5E7EB', fill_type='solid'),  # Gray
}
changed_fill = PatternFill(start_color='FCA5A5', end_color='FCA5A5', fill_type='solid')  # Red for mismatches

# ══════════════════════════════════════════════════════════════════════════
#  Sheet 1: All Pitches (Full Classification)
# ══════════════════════════════════════════════════════════════════════════
ws = wb.active
ws.title = 'All Pitches'

headers = [
    'Pitcher', 'Team', 'Hand', 'Current Label', 'Subtype', 'Suggested Base',
    'Match?', 'Pitches', 'Velo', 'Spin', 'IVB', 'HB',
    'Tilt', 'Distance', '2nd Closest', '2nd Base', '2nd Dist'
]

for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:Q{len(results) + 1}'

for ri, r in enumerate(results, 2):
    match_str = '✓' if not r['changed'] else '✗ MISMATCH'

    vals = [
        r['pitcher'], r['team'], r['hand'], r['current_type'],
        r['subtype'], r['suggested_base'], match_str,
        r['count'],
        round(r['velocity'], 1) if r['velocity'] else '',
        round(r['spin'], 0) if r['spin'] else '',
        round(r['ivb'], 1), round(r['hb'], 1),
        r['tilt'], r['distance'],
        r['second_subtype'], r['second_base'], r['second_distance'],
    ]

    row_fill = changed_fill if r['changed'] else BASE_COLORS.get(r['suggested_base'])

    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=ri, column=ci, value=v)
        cell.font = data_font
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        if ci in (1,):
            cell.alignment = left_align
        else:
            cell.alignment = center_align

    # Bold the mismatch indicator
    if r['changed']:
        ws.cell(row=ri, column=7).font = Font(name='Calibri', size=10, bold=True, color='DC2626')

col_widths = [24, 6, 6, 12, 20, 14, 14, 8, 7, 8, 7, 7, 8, 8, 20, 10, 8]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w


# ══════════════════════════════════════════════════════════════════════════
#  Sheet 2: Mismatches Only
# ══════════════════════════════════════════════════════════════════════════
ws2 = wb.create_sheet('Mismatches')

mismatches = [r for r in results if r['changed']]
mismatches.sort(key=lambda r: (r['current_type'], r['suggested_base'], r['pitcher']))

headers2 = [
    'Pitcher', 'Team', 'Hand', 'Current Label', 'Subtype', 'Suggested Base',
    'Pitches', 'Velo', 'IVB', 'HB', 'Distance', '2nd Closest', '2nd Dist'
]

for ci, h in enumerate(headers2, 1):
    cell = ws2.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f'A1:M{len(mismatches) + 1}'

for ri, r in enumerate(mismatches, 2):
    vals2 = [
        r['pitcher'], r['team'], r['hand'], r['current_type'],
        r['subtype'], r['suggested_base'],
        r['count'],
        round(r['velocity'], 1) if r['velocity'] else '',
        round(r['ivb'], 1), round(r['hb'], 1),
        r['distance'], r['second_subtype'], r['second_distance'],
    ]

    for ci, v in enumerate(vals2, 1):
        cell = ws2.cell(row=ri, column=ci, value=v)
        cell.font = data_font
        cell.border = thin_border
        cell.fill = changed_fill
        cell.alignment = center_align if ci > 1 else left_align

col_widths2 = [24, 6, 6, 12, 20, 14, 8, 7, 7, 7, 8, 20, 8]
for ci, w in enumerate(col_widths2, 1):
    ws2.column_dimensions[get_column_letter(ci)].width = w


# ══════════════════════════════════════════════════════════════════════════
#  Sheet 3: Pitcher Arsenal Summary
# ══════════════════════════════════════════════════════════════════════════
ws3 = wb.create_sheet('Arsenal Summary')

# Group by pitcher
arsenals = defaultdict(list)
for r in results:
    key = (r['pitcher'], r['team'], r['hand'])
    arsenals[key].append(r)

headers3 = [
    'Pitcher', 'Team', 'Hand', 'Pitch', 'Subtype', 'Suggested Base',
    'Match?', 'Pitches', 'Velo', 'IVB', 'HB', 'Distance'
]

for ci, h in enumerate(headers3, 1):
    cell = ws3.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws3.freeze_panes = 'A2'

row_num = 2
separator_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')

for (pitcher, team, hand), pitches in sorted(arsenals.items()):
    # Sort pitches within arsenal: FF, SI, FC, SL, ST, CU, SV
    type_order = {'FF': 0, 'SI': 1, 'FC': 2, 'SL': 3, 'ST': 4, 'CU': 5, 'SV': 6}
    pitches.sort(key=lambda p: type_order.get(p['current_type'], 99))

    for pi, r in enumerate(pitches):
        match_str = '✓' if not r['changed'] else '✗'
        vals3 = [
            pitcher if pi == 0 else '',
            team if pi == 0 else '',
            hand if pi == 0 else '',
            r['current_type'], r['subtype'], r['suggested_base'], match_str,
            r['count'],
            round(r['velocity'], 1) if r['velocity'] else '',
            round(r['ivb'], 1), round(r['hb'], 1), r['distance'],
        ]

        row_fill = changed_fill if r['changed'] else BASE_COLORS.get(r['suggested_base'])

        for ci, v in enumerate(vals3, 1):
            cell = ws3.cell(row=row_num, column=ci, value=v)
            cell.font = bold_font if pi == 0 and ci <= 3 else data_font
            cell.border = thin_border
            if row_fill and ci >= 4:
                cell.fill = row_fill
            cell.alignment = center_align if ci > 1 else left_align

        row_num += 1

    # Add thin separator after each pitcher
    row_num += 0  # No gap needed, the bold name serves as separator

col_widths3 = [24, 6, 6, 8, 20, 14, 8, 8, 7, 7, 7, 8]
for ci, w in enumerate(col_widths3, 1):
    ws3.column_dimensions[get_column_letter(ci)].width = w

ws3.auto_filter.ref = f'A1:L{row_num}'


# ══════════════════════════════════════════════════════════════════════════
#  Sheet 4: Subtype Distribution
# ══════════════════════════════════════════════════════════════════════════
ws4 = wb.create_sheet('Subtype Distribution')

ws4.cell(row=1, column=1, value='Subtype Distribution Summary').font = Font(bold=True, size=14)

headers4 = ['Subtype', 'Base Type', 'Count', 'Avg IVB', 'Avg HB', 'Ref IVB', 'Ref HB']
for ci, h in enumerate(headers4, 1):
    cell = ws4.cell(row=3, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

# Group results by subtype
by_subtype = defaultdict(list)
for r in results:
    by_subtype[r['subtype']].append(r)

# Sort subtypes by base type order then by count
type_order = {'FF': 0, 'SI': 1, 'FC': 2, 'SL': 3, 'ST': 4, 'CU': 5, 'SV': 6}
sorted_subtypes = sorted(by_subtype.items(),
    key=lambda x: (type_order.get(next(b for n, b, _, _ in SUBTYPES if n == x[0]), 99), -len(x[1])))

row4 = 4
for subtype_name, rows in sorted_subtypes:
    base = next(b for n, b, _, _ in SUBTYPES if n == subtype_name)
    ref_ivb = next(iv for n, _, iv, _ in SUBTYPES if n == subtype_name)
    ref_hb = next(hb for n, _, _, hb in SUBTYPES if n == subtype_name)

    # Normalize HB for averaging (flip LHP back)
    norm_hbs = []
    for r in rows:
        nh = r['hb'] if r['hand'] == 'R' else -r['hb']
        norm_hbs.append(nh)

    avg_ivb = sum(r['ivb'] for r in rows) / len(rows)
    avg_hb = sum(norm_hbs) / len(norm_hbs)

    vals4 = [
        subtype_name, base, len(rows),
        round(avg_ivb, 1), round(avg_hb, 1),
        ref_ivb, ref_hb
    ]

    row_fill = BASE_COLORS.get(base)

    for ci, v in enumerate(vals4, 1):
        cell = ws4.cell(row=row4, column=ci, value=v)
        cell.font = data_font
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        cell.alignment = center_align if ci > 1 else left_align

    row4 += 1

col_widths4 = [22, 10, 8, 10, 10, 10, 10]
for ci, w in enumerate(col_widths4, 1):
    ws4.column_dimensions[get_column_letter(ci)].width = w


# ══════════════════════════════════════════════════════════════════════════
#  Sheet 5: Reference Chart
# ══════════════════════════════════════════════════════════════════════════
ws5 = wb.create_sheet('Reference Profiles')

ws5.cell(row=1, column=1, value='Subtype Reference Movement Profiles (RHP Perspective)').font = Font(bold=True, size=14)

headers5 = ['Subtype', 'Base Type', 'Ref IVB (")', 'Ref HB (")', 'Description']
for ci, h in enumerate(headers5, 1):
    cell = ws5.cell(row=3, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

DESCRIPTIONS = {
    'Gyro Fastball': 'Pure gyro spin, minimal movement in either direction',
    'Inefficient FF': 'Below-average ride with moderate arm-side run',
    'Deadzone FB': 'Moderate rise with significant arm-side run — FB/SI border',
    'Running Fastball': 'Good rise with extreme arm-side run',
    'Relative Cut FF': 'Strong ride with minimal horizontal — cuts relative to avg FF',
    'Standard FF': 'Typical four-seam profile with good ride and moderate run',
    'Rider': 'Elite ride with moderate arm-side run',
    'Ride n\' Run': 'Elite ride with significant arm-side run',
    'Gyro Sinker': 'Gyro spin with arm-side run, limited rise',
    'Sinker': 'Classic sinker — low rise, heavy arm-side run',
    'Running Sinker': 'Heavy arm-side run with moderate sink',
    'Heavy Sinker': 'Strong downward action with arm-side run',
    'Heavy Runner': 'Extreme arm-side run with heavy sink',
    'Diver': 'Negative IVB (true drop) with extreme arm-side run',
    'Gyro Cutter': 'Gyro spin, neutral horizontal — true "cutter" action',
    'Standard Cutter': 'Moderate ride with slight glove-side break',
    'Sweeping Cutter': 'Good ride with significant glove-side sweep',
    'Backspinner': 'High ride, near-neutral horizontal — almost a riding fastball shape',
    'Gyro SL': 'Gyro spin with minimal horizontal break',
    'Slutter': 'Slider with cutter-like IVB — "slutter" hybrid',
    'Standard SL': 'Typical slider with moderate glove-side break',
    'Sweeper': 'Extreme glove-side sweep with slight negative IVB',
    'Gyro CB': 'Gyro curveball — moderate drop, minimal horizontal',
    'IE Downer': 'Inefficient downer — good drop, minimal horizontal',
    'Standard CB': 'Classic curveball with strong drop and moderate sweep',
    'Downer': 'Extreme vertical drop with limited horizontal',
    'Efficient CB': 'Maximum drop and significant sweep',
    'IE Slurve': 'Inefficient slurve — moderate drop with moderate sweep',
    'Slurve': 'Drop + sweep hybrid — between curveball and sweeper',
    'Efficient Slurve': 'Maximum sweep with significant drop',
}

row5 = 4
current_base = None
for name, base, ref_ivb, ref_hb in SUBTYPES:
    # Add separator between base type groups
    if base != current_base:
        if current_base is not None:
            row5 += 1  # blank row separator
        current_base = base

    row_fill = BASE_COLORS.get(base)
    vals5 = [name, base, ref_ivb, ref_hb, DESCRIPTIONS.get(name, '')]

    for ci, v in enumerate(vals5, 1):
        cell = ws5.cell(row=row5, column=ci, value=v)
        cell.font = data_font
        cell.border = thin_border
        if row_fill:
            cell.fill = row_fill
        cell.alignment = left_align if ci in (1, 5) else center_align

    row5 += 1

col_widths5 = [22, 10, 10, 10, 60]
for ci, w in enumerate(col_widths5, 1):
    ws5.column_dimensions[get_column_letter(ci)].width = w


# ── Save ──────────────────────────────────────────────────────────────────
output_path = os.path.join(os.path.expanduser('~/Downloads'), 'Pitch_Subtype_Classification.xlsx')
wb.save(output_path)
print(f"\nSaved to: {output_path}")
