#!/usr/bin/env python3
"""
Read outlier list from user's spreadsheet, find matching pitches in Google Sheets,
and blank ONLY the specific metric cell for each outlier.
"""

import gspread
from google.oauth2.service_account import Credentials
import openpyxl
import os
import time as time_module
from collections import defaultdict
from datetime import datetime

SPREADSHEET_ID = '1hNILKCGBuyQKV6KPWawgkS1cu72672TBALi8iNBbIFo'
SERVICE_ACCOUNT_FILE = os.path.join(os.path.dirname(__file__), 'service_account.json')
OUTLIER_FILE = os.path.join(os.path.expanduser('~/Downloads'), 'Untitled spreadsheet (1).xlsx')

MLB_TEAMS = {
    'ARI', 'ATH', 'ATL', 'BAL', 'BOS', 'CHC', 'CIN', 'CLE', 'COL', 'CWS',
    'DET', 'HOU', 'KCR', 'LAA', 'LAD', 'MIA', 'MIL', 'MIN', 'NYM', 'NYY',
    'PHI', 'PIT', 'SDP', 'SEA', 'SFG', 'STL', 'TBR', 'TEX', 'TOR', 'WSH',
}

# Map sheet name in user's xlsx -> column to blank in Google Sheets
SHEET_TO_COL = {
    'Spin': 'Spin Rate',
    'Ext': 'Extension',
    'RelZ': 'RelPosZ',
    'RelX': 'RelPosX',
}


def parse_date(val):
    """Extract YYYY-MM-DD string from various date formats."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if 'T' in s:
        s = s.split('T')[0]
    # Try to extract YYYY-MM-DD
    for fmt in ('%Y-%m-%d', '%m/%d/%Y', '%m/%d/%y'):
        try:
            return datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d')
        except ValueError:
            continue
    return s[:10]


def safe_float(val):
    if val is None or val == '':
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None


def read_outliers():
    """Read the user's outlier spreadsheet. Returns dict: team -> list of (pitcher, pitch_type, date, value, metric_col)."""
    wb = openpyxl.load_workbook(OUTLIER_FILE)
    outliers_by_team = defaultdict(list)
    total = 0

    for sheet_name, target_col in SHEET_TO_COL.items():
        ws = wb[sheet_name]
        # Header row
        headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        pitcher_idx = headers.index('Pitcher')
        team_idx = headers.index('Team')
        pt_idx = headers.index('Pitch Type')
        val_idx = headers.index('Value')
        date_idx = headers.index('Game Date')

        for row in ws.iter_rows(min_row=2, values_only=True):
            pitcher = row[pitcher_idx]
            if not pitcher:
                continue
            team = row[team_idx]
            pitch_type = row[pt_idx]
            value = safe_float(row[val_idx])
            game_date = parse_date(row[date_idx])

            outliers_by_team[team].append({
                'pitcher': pitcher,
                'pitch_type': pitch_type,
                'date': game_date,
                'value': value,
                'target_col': target_col,
                'sheet': sheet_name,
            })
            total += 1

    print(f"Read {total} outliers across {len(outliers_by_team)} teams")
    for team in sorted(outliers_by_team):
        print(f"  {team}: {len(outliers_by_team[team])} outliers")
    return outliers_by_team


def match_value(sheet_val, outlier_val, target_col):
    """Check if a cell value matches the outlier value (with float tolerance)."""
    sv = safe_float(sheet_val)
    if sv is None or outlier_val is None:
        return False
    # For spin rate, compare as integers
    if target_col == 'Spin Rate':
        return abs(round(sv) - round(outlier_val)) < 1
    # For others, compare with small tolerance
    return abs(sv - outlier_val) < 0.015


def match_date(sheet_date_str, outlier_date):
    """Check if dates match."""
    if not sheet_date_str or not outlier_date:
        return False
    sd = parse_date(sheet_date_str)
    return sd == outlier_date


def read_sheet_with_retry(ws, max_retries=3):
    for attempt in range(max_retries):
        try:
            return ws.get_all_values()
        except gspread.exceptions.APIError as e:
            if '429' in str(e) and attempt < max_retries - 1:
                wait = 30 * (attempt + 1)
                print(f"    Rate limited, waiting {wait}s...")
                time_module.sleep(wait)
            else:
                raise


def main():
    outliers_by_team = read_outliers()

    print("\nConnecting to Google Sheets...")
    scopes = ['https://www.googleapis.com/auth/spreadsheets']
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=scopes)
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(SPREADSHEET_ID)
    print(f"Spreadsheet: {sh.title}")

    total_blanked = 0
    total_missed = 0
    missed_details = []

    worksheets = {ws.title: ws for ws in sh.worksheets()}

    for team in sorted(outliers_by_team):
        if team not in worksheets:
            print(f"\n  WARNING: Team {team} not found in spreadsheet!")
            missed_details.extend([(o, 'team not found') for o in outliers_by_team[team]])
            total_missed += len(outliers_by_team[team])
            continue

        ws = worksheets[team]
        team_outliers = outliers_by_team[team]
        print(f"\n  Processing {team} ({len(team_outliers)} outliers)...")

        time_module.sleep(1.5)
        rows = read_sheet_with_retry(ws)
        if not rows:
            print(f"    Empty sheet!")
            continue

        header = rows[0]
        col_idx = {name: i for i, name in enumerate(header) if name}

        # Verify required columns exist
        required = {'Pitcher', 'Pitch Type', 'Game Date', 'Spin Rate', 'Extension', 'RelPosZ', 'RelPosX'}
        missing_cols = required - set(col_idx.keys())
        if missing_cols:
            print(f"    WARNING: Missing columns: {missing_cols}")

        # Build index for faster lookups: (pitcher, pitch_type, date) -> list of (row_idx, row)
        pitch_index = defaultdict(list)
        for row_idx, row in enumerate(rows[1:], start=2):  # 1-indexed for sheets API, skip header
            pitcher = row[col_idx['Pitcher']] if 'Pitcher' in col_idx and col_idx['Pitcher'] < len(row) else ''
            pt = row[col_idx['Pitch Type']] if 'Pitch Type' in col_idx and col_idx['Pitch Type'] < len(row) else ''
            date_raw = row[col_idx['Game Date']] if 'Game Date' in col_idx and col_idx['Game Date'] < len(row) else ''
            date_str = parse_date(date_raw)
            pitch_index[(pitcher, pt, date_str)].append((row_idx, row))

        # Batch updates for this team
        cells_to_blank = []
        team_blanked = 0
        team_missed = 0

        for outlier in team_outliers:
            target_col = outlier['target_col']
            if target_col not in col_idx:
                missed_details.append((outlier, f'column {target_col} not in sheet'))
                team_missed += 1
                continue

            target_col_idx = col_idx[target_col]  # 0-indexed
            target_col_letter = chr(ord('A') + target_col_idx) if target_col_idx < 26 else \
                chr(ord('A') + target_col_idx // 26 - 1) + chr(ord('A') + target_col_idx % 26)

            # Look up matching rows
            key = (outlier['pitcher'], outlier['pitch_type'], outlier['date'])
            candidates = pitch_index.get(key, [])

            matched = False
            for row_idx, row in candidates:
                cell_val = row[target_col_idx] if target_col_idx < len(row) else ''
                if match_value(cell_val, outlier['value'], target_col):
                    # Use gspread column letter conversion
                    col_letter = gspread.utils.rowcol_to_a1(1, target_col_idx + 1)[:-1]
                    cell_ref = f"{col_letter}{row_idx}"
                    cells_to_blank.append(cell_ref)
                    matched = True
                    team_blanked += 1
                    break

            if not matched:
                team_missed += 1
                missed_details.append((outlier, f'no matching row (candidates: {len(candidates)})'))

        # Execute batch blank
        if cells_to_blank:
            # Batch update in chunks (max ~100 per request to stay within limits)
            CHUNK_SIZE = 100
            for i in range(0, len(cells_to_blank), CHUNK_SIZE):
                chunk = cells_to_blank[i:i + CHUNK_SIZE]
                # Use batch_update with list of dicts
                updates = [{'range': cell, 'values': [['']] } for cell in chunk]
                ws.batch_update(updates, value_input_option='RAW')
                if i + CHUNK_SIZE < len(cells_to_blank):
                    time_module.sleep(1)

            print(f"    Blanked {team_blanked} cells, missed {team_missed}")
        else:
            print(f"    No matches found! Missed {team_missed}")

        total_blanked += team_blanked
        total_missed += team_missed

        # Rate limit between teams
        time_module.sleep(1)

    print(f"\n{'='*60}")
    print(f"DONE: Blanked {total_blanked} cells, missed {total_missed}")

    if missed_details:
        print(f"\nMissed outliers ({len(missed_details)}):")
        for outlier, reason in missed_details[:30]:
            print(f"  {outlier['pitcher']} ({outlier['pitch_type']}) {outlier['date']} "
                  f"val={outlier['value']} col={outlier['target_col']} -> {reason}")
        if len(missed_details) > 30:
            print(f"  ... and {len(missed_details) - 30} more")


if __name__ == '__main__':
    main()
