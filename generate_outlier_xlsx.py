#!/usr/bin/env python3
"""Generate Excel outlier report from outlier_results.json."""

import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INPUT = os.path.join(os.path.dirname(__file__), 'outlier_results.json')
OUTPUT = os.path.join(os.path.dirname(__file__), 'ST_2026_Outlier_Report3.xlsx')

with open(INPUT) as f:
    data = json.load(f)

wb = openpyxl.Workbook()
# Remove default sheet
wb.remove(wb.active)

# Styles
header_font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
header_fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
conf_header_fill = PatternFill(start_color='C00000', end_color='C00000', fill_type='solid')
quest_header_fill = PatternFill(start_color='BF8F00', end_color='BF8F00', fill_type='solid')
alt_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
conf_light = PatternFill(start_color='FCE4EC', end_color='FCE4EC', fill_type='solid')
quest_light = PatternFill(start_color='FFF8E1', end_color='FFF8E1', fill_type='solid')
thin_border = Border(
    left=Side(style='thin', color='CCCCCC'),
    right=Side(style='thin', color='CCCCCC'),
    top=Side(style='thin', color='CCCCCC'),
    bottom=Side(style='thin', color='CCCCCC'),
)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center')

HEADERS = ['Pitcher', 'Team', 'Pitch Type', 'Value', 'Pitcher Avg', 'Diff', 'Std Dev', 'Z-Score', '# Pitches', 'Game Date']
# Column widths
COL_WIDTHS = [22, 7, 10, 14, 14, 12, 10, 9, 10, 12]

SECTIONS = [
    ('Spin Rate', 'Spin Rate'),
    ('Extension', 'Extension'),
    ('Release Height', 'Release Height'),
    ('Release Side', 'Release Side'),
]


def format_val(val, unit, rnd):
    if unit == 'RPM':
        return int(round(val))
    else:
        return round(val, 2)


def write_section(ws, start_row, records, section_fill, label, metric_info):
    """Write a section (confident or questionable) and return next row."""
    unit = metric_info['unit']
    rnd = metric_info['round']

    if not records:
        return start_row

    # Section label row
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(HEADERS))
    cell = ws.cell(row=start_row, column=1, value=f"{label} ({len(records)})")
    cell.font = Font(name='Arial', bold=True, size=11, color='FFFFFF')
    cell.fill = section_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    for c in range(2, len(HEADERS) + 1):
        ws.cell(row=start_row, column=c).fill = section_fill
    start_row += 1

    # Header row
    for i, h in enumerate(HEADERS, 1):
        cell = ws.cell(row=start_row, column=i, value=h)
        cell.font = Font(name='Arial', bold=True, size=9, color='333333')
        cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
        cell.alignment = center_align
        cell.border = thin_border
    start_row += 1

    # Data rows
    for idx, r in enumerate(records):
        is_alt = idx % 2 == 1
        row_fill = alt_fill if is_alt else PatternFill(fill_type=None)

        date_str = r.get('game_date', '')
        if date_str and 'T' in date_str:
            date_str = date_str.split('T')[0]

        diff = r['value'] - r['mean']
        if unit == 'RPM':
            diff_str = int(round(diff))
        else:
            diff_str = round(diff, 2)

        vals = [
            r['pitcher'],
            r['team'],
            r['pitch_type'],
            format_val(r['value'], unit, rnd),
            format_val(r['mean'], unit, rnd),
            diff_str,
            f"{r['std']:.1f}" if unit == 'RPM' else f"{r['std']:.2f}",
            f"{r['z_score']:.2f}",
            r['n_pitches'],
            date_str,
        ]

        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=start_row, column=i, value=v)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = center_align if i >= 4 else left_align
            cell.border = thin_border
            if is_alt:
                cell.fill = row_fill

        start_row += 1

    return start_row + 1  # blank row after section


METRIC_INFO = {
    'Spin Rate': {'unit': 'RPM', 'round': 0},
    'Extension': {'unit': 'ft', 'round': 2},
    'Release Height': {'unit': 'ft', 'round': 2},
    'Release Side': {'unit': 'ft', 'round': 2},
}

# Create a sheet for each metric
for section_key, sheet_name in SECTIONS:
    ws = wb.create_sheet(title=sheet_name)

    # Set column widths
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Freeze top rows
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

    conf_records = data[section_key]['confident']
    quest_records = data[section_key]['questionable']
    mi = METRIC_INFO[section_key]

    row = 1

    # Title
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    cell = ws.cell(row=row, column=1, value=f"{sheet_name} Outliers")
    cell.font = Font(name='Arial', bold=True, size=14)
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1

    # Subtitle with counts
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(HEADERS))
    cell = ws.cell(row=row, column=1,
                   value=f"{len(conf_records)} confident  |  {len(quest_records)} questionable  |  {len(conf_records)+len(quest_records)} total")
    cell.font = Font(name='Arial', size=10, color='666666')
    row += 2

    # Confident outliers
    row = write_section(ws, row, conf_records, conf_header_fill, 'Confident Outliers', mi)

    # Questionable outliers
    row = write_section(ws, row, quest_records, quest_header_fill, 'Questionable Outliers', mi)

    # Freeze pane at row 4 (after title rows)
    ws.freeze_panes = 'A4'

# Summary sheet
ws = wb.create_sheet(title='Summary', index=0)
ws.column_dimensions['A'].width = 20
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 10

ws.cell(row=1, column=1, value='ST 2026 Data Outlier Report').font = Font(name='Arial', bold=True, size=16)
ws.merge_cells('A1:D1')
ws.cell(row=2, column=1, value='March 12, 2026').font = Font(name='Arial', size=10, color='777777')
ws.merge_cells('A2:D2')

# Methodology
ws.merge_cells('A4:D4')
ws.cell(row=4, column=1, value='Methodology').font = Font(name='Arial', bold=True, size=12)
ws.merge_cells('A5:D8')
ws.cell(row=5, column=1, value=(
    'Outliers detected per-pitcher, per-pitch-type using Z-score (\u22653.0 confident, \u22652.2 questionable) '
    'cross-validated with IQR bounds (2.0x). Minimum 5 pitches per group. '
    'Absolute deviation thresholds: Spin Rate \u2265200 RPM (300 for CH/FS/KN), Extension \u22650.4 ft, '
    'Release Height \u22650.3 ft, Release Side \u22650.3 ft.'
)).font = Font(name='Arial', size=10)
ws.cell(row=5, column=1).alignment = Alignment(wrap_text=True, vertical='top')
ws.row_dimensions[5].height = 60

# Summary table
sum_headers = ['Metric', 'Confident', 'Questionable', 'Total']
for i, h in enumerate(sum_headers, 1):
    cell = ws.cell(row=10, column=i, value=h)
    cell.font = Font(name='Arial', bold=True, size=10, color='FFFFFF')
    cell.fill = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
    cell.alignment = center_align
    cell.border = thin_border

grand_c, grand_q = 0, 0
for idx, (section_key, sheet_name) in enumerate(SECTIONS):
    r = 11 + idx
    c = len(data[section_key]['confident'])
    q = len(data[section_key]['questionable'])
    grand_c += c
    grand_q += q
    is_alt = idx % 2 == 1

    vals = [sheet_name, c, q, c + q]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=r, column=i, value=v)
        cell.font = Font(name='Arial', size=10)
        cell.alignment = center_align if i >= 2 else left_align
        cell.border = thin_border
        if is_alt:
            cell.fill = alt_fill

# Total row
r = 15
vals = ['Total', grand_c, grand_q, grand_c + grand_q]
for i, v in enumerate(vals, 1):
    cell = ws.cell(row=r, column=i, value=v)
    cell.font = Font(name='Arial', bold=True, size=10)
    cell.alignment = center_align if i >= 2 else left_align
    cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')
    cell.border = thin_border

# Add autofilter to each metric sheet
for section_key, sheet_name in SECTIONS:
    sheet = wb[sheet_name]
    # Find the first header row (row with "Pitcher" in column A after the section label)
    for row_num in range(1, sheet.max_row + 1):
        if sheet.cell(row=row_num, column=1).value == 'Pitcher':
            last_col = get_column_letter(len(HEADERS))
            sheet.auto_filter.ref = f"A{row_num}:{last_col}{sheet.max_row}"
            break

wb.save(OUTPUT)
print(f"Saved to {OUTPUT}")
print(f"Sheets: {[s.title for s in wb.worksheets]}")
