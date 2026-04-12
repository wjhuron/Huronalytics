#!/usr/bin/env python3
"""
Pitch Reclassification Analysis v2
Analyzes every pitcher's arsenal for potential pitch type misclassifications.
Focuses on: FF, FC, SL, ST, CU, SV

Classification guidelines:
  - Curveball (CU): IVB <= -6"
  - Sweeper (ST): |HB| >= 10" (glove-side break >= 10")
  - Cutter (FC): IVB >= 6" AND |HB| <= 4" (glove-side break <= 4")
  - Fastball (FF): High IVB (12"+), arm-side run, high velo
  - Slider (SL): Moderate IVB (-6" to 6"), moderate glove-side break (4"-10")
  - Slurve (SV): Negative IVB with significant horizontal break (drop + sweep hybrid)

Conflict rule: A pitcher cannot have two of the same pitch type.
  If suggesting X→Y but pitcher already has Y, only keep the suggestion
  if Y→Z is also suggested (i.e., the existing Y is being reclassified).
"""

import json
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Load data ──────────────────────────────────────────────────────────────
with open('data/pitch_leaderboard.json') as f:
    pitch_data = json.load(f)

TARGET_TYPES = {'FF', 'FC', 'SL', 'ST', 'CU', 'SV'}

# ── League averages by pitch type and hand ─────────────────────────────────
league_avgs = {}
for hand in ('R', 'L'):
    league_avgs[hand] = {}
    for pt in TARGET_TYPES:
        rows = [p for p in pitch_data if p['throws'] == hand and p['pitchType'] == pt
                and p.get('velocity') and p.get('indVertBrk') is not None and p.get('horzBrk') is not None]
        if not rows:
            continue
        league_avgs[hand][pt] = {
            'velo': sum(r['velocity'] for r in rows) / len(rows),
            'ivb': sum(r['indVertBrk'] for r in rows) / len(rows),
            'hb': sum(r['horzBrk'] for r in rows) / len(rows),
            'spin': sum(r['spinRate'] for r in rows if r.get('spinRate')) / max(1, len([r for r in rows if r.get('spinRate')])),
            'n': len(rows),
        }

print("League averages (RHP):")
for pt in ['FF', 'FC', 'SL', 'ST', 'CU', 'SV']:
    if pt in league_avgs['R']:
        a = league_avgs['R'][pt]
        print(f"  {pt}: Velo={a['velo']:.1f}, IVB={a['ivb']:.1f}, HB={a['hb']:.1f}, Spin={a['spin']:.0f} (n={a['n']})")

print("\nLeague averages (LHP):")
for pt in ['FF', 'FC', 'SL', 'ST', 'CU', 'SV']:
    if pt in league_avgs['L']:
        a = league_avgs['L'][pt]
        print(f"  {pt}: Velo={a['velo']:.1f}, IVB={a['ivb']:.1f}, HB={a['hb']:.1f}, Spin={a['spin']:.0f} (n={a['n']})")


# ── Build pitcher arsenals ─────────────────────────────────────────────────
arsenals = defaultdict(dict)
for p in pitch_data:
    if p['pitchType'] not in TARGET_TYPES:
        continue
    if not p.get('velocity') or p.get('indVertBrk') is None or p.get('horzBrk') is None:
        continue
    key = (p['pitcher'], p['team'], p['throws'])
    arsenals[key][p['pitchType']] = {
        'velo': p['velocity'],
        'ivb': p['indVertBrk'],
        'hb': p['horzBrk'],
        'spin': p.get('spinRate', 0),
        'tilt': p.get('breakTilt', ''),
        'vaa': p.get('vaa'),
        'count': p.get('count', 0),
    }


# ── Helper functions ──────────────────────────────────────────────────────
def glove_side_break(hb, hand):
    """Positive = more glove-side movement."""
    return -hb if hand == 'R' else hb

def arm_side_run(hb, hand):
    """Positive = more arm-side movement."""
    return hb if hand == 'R' else -hb


# ── Classification logic ──────────────────────────────────────────────────
def analyze_pitch(pitcher_key, pt, metrics, arsenal):
    """
    Analyze a single pitch type for potential reclassification.
    Returns list of (suggested_type, confidence, reason).
    """
    hand = pitcher_key[2]
    velo = metrics['velo']
    ivb = metrics['ivb']
    hb = metrics['hb']
    spin = metrics['spin']
    count = metrics['count']
    gsb = glove_side_break(hb, hand)  # positive = glove-side
    asr = arm_side_run(hb, hand)      # positive = arm-side

    # Pitcher's FF for context
    ff = arsenal.get('FF')
    ff_velo = ff['velo'] if ff else None

    suggestions = []

    # ══════════════════════════════════════════════════════════════════
    #  FF analysis
    # ══════════════════════════════════════════════════════════════════
    if pt == 'FF':
        # FF → FC: Cutter profile = IVB < 10, glove-side break, often lower velo
        # Strong case: IVB well below FF avg (~15), glove-side break
        if ivb < 8 and gsb > 2:
            suggestions.append(('FC', 'High',
                f"Low IVB ({ivb:.1f}\") with glove-side break ({gsb:.1f}\") at {velo:.1f} mph. "
                f"Cutters typically have IVB >= 6\" with glove-side break <= 4\". "
                f"League avg FF IVB is ~15\" — this is far below."))
        elif ivb < 10 and gsb > 0:
            suggestions.append(('FC', 'Medium',
                f"IVB ({ivb:.1f}\") is below typical FF range (~15\") with some glove-side break ({gsb:.1f}\"). "
                f"This could be a cutter rather than a four-seam."))

    # ══════════════════════════════════════════════════════════════════
    #  FC analysis — Cutter norms: IVB >= 6", GSB <= 4"
    # ══════════════════════════════════════════════════════════════════
    elif pt == 'FC':
        # FC → FF: High IVB with arm-side run = fastball
        if ivb > 12 and asr > 3:
            suggestions.append(('FF', 'High',
                f"High IVB ({ivb:.1f}\") with arm-side run ({asr:.1f}\") at {velo:.1f} mph is a fastball profile. "
                f"Cutters have IVB ~6-10\" with glove-side break — this rides like a four-seam."))
        elif ivb > 10 and asr > 1:
            suggestions.append(('FF', 'Medium',
                f"IVB ({ivb:.1f}\") and arm-side movement ({asr:.1f}\") trend toward fastball territory. "
                f"Cutters typically have IVB 6-10\" with neutral-to-glove-side break."))

        # FC → SL: Low IVB + significant glove-side break = slider
        # Below cutter norms: IVB < 6 or GSB > 4
        if ivb < 3 and gsb > 4:
            suggestions.append(('SL', 'High',
                f"Very low IVB ({ivb:.1f}\") with significant glove-side break ({gsb:.1f}\") at {velo:.1f} mph. "
                f"Cutters should have IVB >= 6\" and GSB <= 4\" — this is slider shape."))
        elif ivb < 5 and gsb > 4:
            suggestions.append(('SL', 'Medium',
                f"IVB ({ivb:.1f}\") below cutter norm (>= 6\") with glove-side break ({gsb:.1f}\") exceeding "
                f"cutter range (<= 4\"). This is on the SL side of the FC/SL spectrum."))

    # ══════════════════════════════════════════════════════════════════
    #  SL analysis — Slider: IVB roughly -6" to 6", GSB 4"-10"
    # ══════════════════════════════════════════════════════════════════
    elif pt == 'SL':
        # SL → ST: High glove-side break = sweeper (GSB >= 10")
        if gsb >= 12:
            suggestions.append(('ST', 'High',
                f"Glove-side break of {gsb:.1f}\" is firmly in sweeper territory (>= 10\"). "
                f"League avg SL GSB is ~4\", ST is ~13\". This sweeps like a sweeper."))
        elif gsb >= 10:
            suggestions.append(('ST', 'Medium',
                f"Glove-side break of {gsb:.1f}\" is at the sweeper threshold (>= 10\"). "
                f"This is borderline SL/ST with sweeper-level horizontal movement."))

        # SL → FC: High IVB + low glove-side break + higher velo = cutter
        # Cutter norms: IVB >= 6, GSB <= 4
        if ivb >= 7 and gsb <= 3:
            suggestions.append(('FC', 'High',
                f"High IVB ({ivb:.1f}\") with minimal glove-side break ({gsb:.1f}\") at {velo:.1f} mph. "
                f"Cutters have IVB >= 6\" and GSB <= 4\" — this fits cutter shape perfectly."))
        elif ivb >= 6 and gsb <= 4:
            suggestions.append(('FC', 'Medium',
                f"IVB ({ivb:.1f}\") and limited horizontal break ({gsb:.1f}\") at {velo:.1f} mph "
                f"match cutter norms (IVB >= 6\", GSB <= 4\"). Borderline FC/SL."))

        # SL → CU: Very negative IVB = curveball (IVB <= -6")
        if ivb <= -6:
            suggestions.append(('CU', 'High',
                f"IVB of {ivb:.1f}\" is in curveball territory (<= -6\"). "
                f"League avg SL IVB is ~2\", CU is ~-10\". This drops like a curve."))
        elif ivb < -4:
            suggestions.append(('CU', 'Medium',
                f"IVB of {ivb:.1f}\" is trending toward curveball shape (CU threshold: <= -6\"). "
                f"Well below typical slider IVB (~2\")."))

        # SL → SV: Negative IVB + high glove-side break = slurve (drop + sweep)
        if ivb < -3 and gsb > 8:
            suggestions.append(('SV', 'High',
                f"IVB ({ivb:.1f}\") with glove-side break ({gsb:.1f}\") is a slurve — "
                f"drop + sweep hybrid. League avg SV: IVB ~-6\", GSB ~11\"."))
        elif ivb < -2 and gsb > 6:
            suggestions.append(('SV', 'Medium',
                f"IVB ({ivb:.1f}\") combined with glove-side break ({gsb:.1f}\") fits a slurve profile. "
                f"More drop than a typical slider with meaningful sweep."))

    # ══════════════════════════════════════════════════════════════════
    #  ST analysis — Sweeper norms: GSB >= 10"
    # ══════════════════════════════════════════════════════════════════
    elif pt == 'ST':
        # ST → SL: Insufficient glove-side break for sweeper
        if gsb < 7:
            suggestions.append(('SL', 'High',
                f"Glove-side break of only {gsb:.1f}\" is well below sweeper threshold (>= 10\"). "
                f"League avg ST GSB is ~13\". This is a traditional slider."))
        elif gsb < 10:
            suggestions.append(('SL', 'Medium',
                f"Glove-side break of {gsb:.1f}\" is below the sweeper threshold (>= 10\"). "
                f"Borderline — could be classified as a slider with extra sweep."))

        # ST → SV: Negative IVB + sweep = slurve (dropping sweeper)
        if ivb < -4 and gsb >= 6:
            suggestions.append(('SV', 'High',
                f"IVB of {ivb:.1f}\" with sweep ({gsb:.1f}\") is a slurve — "
                f"too much drop for a sweeper (league avg ST IVB ~1\"). Drop + sweep = slurve."))
        elif ivb < -3 and gsb >= 6:
            suggestions.append(('SV', 'Medium',
                f"IVB of {ivb:.1f}\" with sweep ({gsb:.1f}\") trends toward slurve. "
                f"Sweepers typically have IVB around 1\" — this has more drop than expected."))

        # ST → CU: Very negative IVB
        if ivb <= -6:
            suggestions.append(('CU', 'High',
                f"IVB of {ivb:.1f}\" is curveball territory (<= -6\"), regardless of sweep ({gsb:.1f}\"). "
                f"This pitch drops too much to be a sweeper."))

        # ST → FC: High IVB, low sweep = cutter shape
        if ivb >= 6 and gsb < 5:
            suggestions.append(('FC', 'Medium',
                f"IVB ({ivb:.1f}\") with limited sweep ({gsb:.1f}\") at {velo:.1f} mph "
                f"is more cutter than sweeper."))

    # ══════════════════════════════════════════════════════════════════
    #  CU analysis — Curveball norms: IVB <= -6"
    # ══════════════════════════════════════════════════════════════════
    elif pt == 'CU':
        # CU → SL: IVB too high for a curve (not enough drop)
        if ivb > 0:
            suggestions.append(('SL', 'High',
                f"Positive IVB ({ivb:.1f}\") — this pitch doesn't drop at all. "
                f"CU needs IVB <= -6\". This is slider shape."))
        elif ivb > -3:
            suggestions.append(('SL', 'High',
                f"IVB of {ivb:.1f}\" is far above curveball range (<= -6\"). "
                f"League avg CU IVB is ~-10\". This is a slider."))
        elif ivb > -6:
            suggestions.append(('SL', 'Medium',
                f"IVB of {ivb:.1f}\" is above the curveball threshold (<= -6\"). "
                f"Borderline — not quite enough drop for a true curveball."))

        # CU → SV: Moderate drop + lots of horizontal sweep = slurve
        if ivb > -8 and gsb > 10:
            suggestions.append(('SV', 'Medium',
                f"IVB ({ivb:.1f}\") is shallower than typical CU (~-10\") with significant sweep ({gsb:.1f}\"). "
                f"This drop + sweep hybrid fits a slurve profile."))
        # Stronger case: clearly between CU and ST in shape
        if ivb > -6 and gsb > 10:
            suggestions.append(('SV', 'High',
                f"IVB ({ivb:.1f}\") above CU threshold with heavy sweep ({gsb:.1f}\"). "
                f"Not enough drop for CU, too much drop for ST — this is a slurve."))

        # CU → ST: Very little drop + lots of sweep = sweeper
        if ivb > -3 and gsb >= 10:
            suggestions.append(('ST', 'High',
                f"Minimal drop ({ivb:.1f}\" IVB) with heavy sweep ({gsb:.1f}\") is a sweeper. "
                f"Curveballs need IVB <= -6\" — this has sweeper shape."))

    # ══════════════════════════════════════════════════════════════════
    #  SV analysis — Slurve: negative IVB + high GSB (drop + sweep)
    # ══════════════════════════════════════════════════════════════════
    elif pt == 'SV':
        # SV → CU: Lots of drop, less horizontal = curveball
        if ivb <= -8 and gsb < 8:
            suggestions.append(('CU', 'High',
                f"Deep drop ({ivb:.1f}\" IVB) with moderate horizontal ({gsb:.1f}\") is curveball shape. "
                f"The vertical drop dominates — this is a curve, not a slurve."))
        elif ivb <= -8:
            suggestions.append(('CU', 'Medium',
                f"IVB of {ivb:.1f}\" is deep in curveball range, even with sweep ({gsb:.1f}\")."))

        # SV → ST: Very little drop + lots of sweep = sweeper
        if ivb > 0 and gsb >= 8:
            suggestions.append(('ST', 'High',
                f"Positive IVB ({ivb:.1f}\") with sweep ({gsb:.1f}\") — no drop at all. "
                f"This is a sweeper, not a slurve. Slurves need negative IVB."))
        elif ivb > -2 and gsb >= 10:
            suggestions.append(('ST', 'Medium',
                f"Minimal drop ({ivb:.1f}\" IVB) with significant sweep ({gsb:.1f}\") trends sweeper."))

        # SV → SL: Low sweep for a slurve, moderate drop
        if gsb < 5 and ivb > -6:
            suggestions.append(('SL', 'High',
                f"Limited sweep ({gsb:.1f}\") and moderate drop ({ivb:.1f}\") is slider shape. "
                f"Slurves need significant horizontal break (~10\"+) with drop."))
        elif gsb < 7 and ivb > -5:
            suggestions.append(('SL', 'Medium',
                f"Sweep ({gsb:.1f}\") and drop ({ivb:.1f}\") are both moderate — "
                f"this is more slider than slurve."))

        # SV → FC: High IVB + low sweep = cutter
        if ivb >= 4 and gsb < 5:
            suggestions.append(('FC', 'High',
                f"IVB ({ivb:.1f}\") with limited break ({gsb:.1f}\") at {velo:.1f} mph. "
                f"This has cutter shape, not slurve."))

    return suggestions


# ── Run analysis (pass 1: generate all suggestions) ──────────────────────
raw_results = []

for pitcher_key, arsenal in arsenals.items():
    pitcher_name, team, hand = pitcher_key

    for pt, metrics in arsenal.items():
        suggestions = analyze_pitch(pitcher_key, pt, metrics, arsenal)

        for suggested_type, confidence, reason in suggestions:
            # Build arsenal context notes
            notes = []
            ff = arsenal.get('FF')
            if ff and pt != 'FF':
                notes.append(f"FF: {ff['velo']:.1f} mph, {ff['ivb']:.1f}\" IVB")
            fc = arsenal.get('FC')
            if fc and pt != 'FC':
                notes.append(f"FC: {fc['velo']:.1f} mph, {fc['ivb']:.1f}\" IVB, {fc['hb']:.1f}\" HB")

            # Velocity gap for FF/FC suggestions
            if (pt == 'FF' and suggested_type == 'FC') or (pt == 'FC' and suggested_type == 'FF'):
                ff_m = arsenal.get('FF')
                fc_m = arsenal.get('FC')
                if ff_m and fc_m:
                    gap = ff_m['velo'] - fc_m['velo']
                    notes.append(f"FF-FC velo gap: {gap:.1f} mph")

            has_suggested = suggested_type in arsenal

            raw_results.append({
                'pitcher': pitcher_name,
                'team': team,
                'hand': hand,
                'current_type': pt,
                'suggested_type': suggested_type,
                'confidence': confidence,
                'count': metrics['count'],
                'velo': metrics['velo'],
                'spin': metrics['spin'],
                'ivb': metrics['ivb'],
                'hb': metrics['hb'],
                'tilt': metrics['tilt'],
                'vaa': metrics['vaa'],
                'reason': reason,
                'arsenal_context': ' | '.join(notes) if notes else '',
                'has_suggested_already': has_suggested,
                'pitcher_key': pitcher_key,
            })


# ── Pass 2: Conflict resolution ──────────────────────────────────────────
# If suggesting X→Y but pitcher already has Y, only keep if Y→Z also exists
# (i.e., the existing Y is being moved out of the way)

# Build index: for each pitcher, what types are being suggested to change?
pitcher_changes = defaultdict(set)  # pitcher_key → set of current_types being changed
for r in raw_results:
    pitcher_changes[r['pitcher_key']].add(r['current_type'])

results = []
removed_conflicts = []

for r in raw_results:
    if r['has_suggested_already']:
        # Pitcher already has the suggested type — check if that type is also being changed
        suggested = r['suggested_type']
        if suggested in pitcher_changes[r['pitcher_key']]:
            # The existing pitch of that type IS being reclassified → keep this suggestion
            r['conflict_note'] = f"Note: Pitcher already has {suggested}, but that {suggested} is also flagged for reclassification."
            results.append(r)
        else:
            # Conflict — remove this suggestion
            removed_conflicts.append(r)
    else:
        r['conflict_note'] = ''
        results.append(r)

# Deduplicate: if a pitch gets multiple suggestions to the same type (e.g., CU→SL from two rules),
# keep only the highest confidence one
seen = {}
deduped = []
for r in results:
    key = (r['pitcher_key'], r['current_type'], r['suggested_type'])
    conf_order = {'High': 0, 'Medium': 1}
    if key not in seen or conf_order.get(r['confidence'], 2) < conf_order.get(seen[key]['confidence'], 2):
        seen[key] = r

deduped = list(seen.values())

# Sort: High confidence first, then by pitcher name
confidence_order = {'High': 0, 'Medium': 1}
deduped.sort(key=lambda r: (confidence_order.get(r['confidence'], 2), r['pitcher'], r['current_type']))

results = deduped

print(f"\nTotal suggestions (after conflict resolution): {len(results)}")
print(f"  High confidence: {len([r for r in results if r['confidence'] == 'High'])}")
print(f"  Medium confidence: {len([r for r in results if r['confidence'] == 'Medium'])}")
print(f"  Removed due to conflicts: {len(removed_conflicts)}")

# Count by change type
from collections import Counter
changes = Counter()
for r in results:
    changes[f"{r['current_type']} → {r['suggested_type']}"] += 1
print("\nChange type breakdown:")
for change, cnt in changes.most_common():
    print(f"  {change}: {cnt}")


# ── Generate Excel ────────────────────────────────────────────────────────
wb = Workbook()

# ── Styles ──
header_font = Font(name='Calibri', bold=True, size=11, color='FFFFFF')
header_fill = PatternFill(start_color='2F3640', end_color='2F3640', fill_type='solid')
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

high_fill = PatternFill(start_color='FFD5D5', end_color='FFD5D5', fill_type='solid')    # Light red
medium_fill = PatternFill(start_color='FFF3CD', end_color='FFF3CD', fill_type='solid')  # Light yellow

data_font = Font(name='Calibri', size=10)
bold_font = Font(name='Calibri', size=10, bold=True)
center_align = Alignment(horizontal='center', vertical='center')
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0'),
)

# ── Sheet 1: All Suggestions ──
ws = wb.active
ws.title = 'Pitch Reclassifications'

headers = [
    'Pitcher', 'Team', 'Hand', 'Current Type', 'Suggested Type', 'Confidence',
    'Pitches', 'Velo (mph)', 'Spin (rpm)', 'IVB (")', 'HB (")', 'Tilt',
    'Reason', 'Arsenal Context', 'Notes'
]

for ci, h in enumerate(headers, 1):
    cell = ws.cell(row=1, column=ci, value=h)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_align
    cell.border = thin_border

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f'A1:O{len(results) + 1}'

for ri, r in enumerate(results, 2):
    notes = r.get('conflict_note', '')
    vals = [
        r['pitcher'], r['team'], r['hand'], r['current_type'], r['suggested_type'],
        r['confidence'], r['count'],
        round(r['velo'], 1), round(r['spin'], 0) if r['spin'] else '',
        round(r['ivb'], 1), round(r['hb'], 1), r['tilt'],
        r['reason'], r['arsenal_context'], notes,
    ]
    row_fill = high_fill if r['confidence'] == 'High' else medium_fill

    for ci, v in enumerate(vals, 1):
        cell = ws.cell(row=ri, column=ci, value=v)
        cell.font = data_font
        cell.border = thin_border
        cell.fill = row_fill
        if ci in (1, 13, 14, 15):  # Pitcher, Reason, Arsenal Context, Notes
            cell.alignment = left_align
        else:
            cell.alignment = center_align

# Column widths
col_widths = [22, 6, 6, 12, 14, 12, 8, 10, 10, 8, 8, 8, 65, 55, 40]
for ci, w in enumerate(col_widths, 1):
    ws.column_dimensions[get_column_letter(ci)].width = w


# ── Sheet 2: Summary ──
ws2 = wb.create_sheet('Summary')

ws2.cell(row=1, column=1, value='Pitch Reclassification Summary').font = Font(bold=True, size=14)
ws2.cell(row=2, column=1, value=f'Total suggestions: {len(results)}').font = Font(size=11)
ws2.cell(row=3, column=1, value=f'High confidence (red): {len([r for r in results if r["confidence"] == "High"])}').font = Font(size=11)
ws2.cell(row=4, column=1, value=f'Medium / borderline (yellow): {len([r for r in results if r["confidence"] == "Medium"])}').font = Font(size=11)
ws2.cell(row=5, column=1, value=f'Removed due to arsenal conflicts: {len(removed_conflicts)}').font = Font(size=11, italic=True)

row = 7
summary_headers = ['Change Type', 'Total', 'High Conf.', 'Medium Conf.']
for ci, h in enumerate(summary_headers, 1):
    cell = ws2.cell(row=row, column=ci, value=h)
    cell.font = Font(bold=True, size=11)
    cell.border = thin_border
    cell.fill = header_fill
    cell.font = header_font

for change, cnt in changes.most_common():
    row += 1
    high_cnt = len([r for r in results if f"{r['current_type']} → {r['suggested_type']}" == change and r['confidence'] == 'High'])
    med_cnt = cnt - high_cnt
    ws2.cell(row=row, column=1, value=change).border = thin_border
    ws2.cell(row=row, column=2, value=cnt).border = thin_border
    ws2.cell(row=row, column=3, value=high_cnt).border = thin_border
    ws2.cell(row=row, column=4, value=med_cnt).border = thin_border
    ws2.cell(row=row, column=2).alignment = center_align
    ws2.cell(row=row, column=3).alignment = center_align
    ws2.cell(row=row, column=4).alignment = center_align

# Classification guidelines reference
row += 2
ws2.cell(row=row, column=1, value='Classification Guidelines Used').font = Font(bold=True, size=12)
row += 1
guidelines = [
    ('Curveball (CU)', 'IVB <= -6"'),
    ('Sweeper (ST)', 'Glove-side break >= 10"'),
    ('Cutter (FC)', 'IVB >= 6" AND glove-side break <= 4"'),
    ('Fastball (FF)', 'High IVB (12"+), arm-side run, high velocity'),
    ('Slider (SL)', 'IVB roughly -6" to 6", glove-side break 4"-10"'),
    ('Slurve (SV)', 'Negative IVB (drop) + significant horizontal break (sweep hybrid)'),
]
for label, desc in guidelines:
    ws2.cell(row=row, column=1, value=label).font = Font(bold=True, size=10)
    ws2.cell(row=row, column=2, value=desc).font = Font(size=10)
    row += 1

ws2.column_dimensions['A'].width = 22
ws2.column_dimensions['B'].width = 55
ws2.column_dimensions['C'].width = 12
ws2.column_dimensions['D'].width = 14


# ── Sheet 3: League Averages Reference ──
ws3 = wb.create_sheet('League Averages')

ws3.cell(row=1, column=1, value='League Average Movement Profiles (Reference)').font = Font(bold=True, size=14)

for hi, hand in enumerate(['R', 'L']):
    start_row = 3 + hi * 10
    ws3.cell(row=start_row, column=1, value=f'{"Right" if hand == "R" else "Left"}-Handed Pitchers').font = Font(bold=True, size=12)

    hdr_row = start_row + 1
    for ci, h in enumerate(['Pitch Type', 'Avg Velo', 'Avg IVB', 'Avg HB', 'Avg Spin', 'Count'], 1):
        cell = ws3.cell(row=hdr_row, column=ci, value=h)
        cell.font = Font(bold=True)
        cell.border = thin_border

    for pi, pt in enumerate(['FF', 'FC', 'SL', 'ST', 'CU', 'SV']):
        r = hdr_row + 1 + pi
        if pt in league_avgs.get(hand, {}):
            a = league_avgs[hand][pt]
            ws3.cell(row=r, column=1, value=pt).border = thin_border
            ws3.cell(row=r, column=2, value=round(a['velo'], 1)).border = thin_border
            ws3.cell(row=r, column=3, value=round(a['ivb'], 1)).border = thin_border
            ws3.cell(row=r, column=4, value=round(a['hb'], 1)).border = thin_border
            ws3.cell(row=r, column=5, value=round(a['spin'], 0)).border = thin_border
            ws3.cell(row=r, column=6, value=a['n']).border = thin_border

for ci in range(1, 7):
    ws3.column_dimensions[get_column_letter(ci)].width = 14


# ── Sheet 4: Removed Conflicts ──
ws4 = wb.create_sheet('Removed (Conflicts)')

ws4.cell(row=1, column=1, value='Suggestions removed because pitcher already has the suggested pitch type').font = Font(bold=True, size=11, italic=True)

if removed_conflicts:
    headers4 = ['Pitcher', 'Team', 'Hand', 'Current Type', 'Suggested Type', 'Confidence',
                 'Reason', 'Already Has']
    for ci, h in enumerate(headers4, 1):
        cell = ws4.cell(row=2, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for ri, r in enumerate(removed_conflicts, 3):
        vals4 = [r['pitcher'], r['team'], r['hand'], r['current_type'], r['suggested_type'],
                 r['confidence'], r['reason'], 'Yes']
        for ci, v in enumerate(vals4, 1):
            cell = ws4.cell(row=ri, column=ci, value=v)
            cell.font = data_font
            cell.border = thin_border

    col_widths4 = [22, 6, 6, 12, 14, 12, 65, 10]
    for ci, w in enumerate(col_widths4, 1):
        ws4.column_dimensions[get_column_letter(ci)].width = w


# ── Save ──────────────────────────────────────────────────────────────────
output_path = os.path.join(os.path.expanduser('~/Downloads'), 'Pitch_Reclassification_Analysis.xlsx')
wb.save(output_path)
print(f"\nSaved to: {output_path}")
