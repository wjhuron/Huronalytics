#!/usr/bin/env python3
"""
Backfill supplemental Statcast data into the Google Sheet.

Scans each team tab for rows that have a PitchID but are missing supplemental
columns (ArmAngle, BatSpeed, SwingLength, AttackAngle,
AttackDirection, SwingPathTilt). Downloads the Statcast Search CSV for the
relevant team/date ranges from Baseball Savant and fills in the empty cells.

Configuration: edit the variables below before running.
"""

import argparse
import gspread
import requests
import pandas as pd
from io import StringIO
import os
import time
from datetime import datetime

# ── USER CONFIGURATION ──────────────────────────────────────────────────────
# Set date range (inclusive). Leave both as None to backfill all dates.
start_date = None
end_date   = None

# Set specific teams, or None for all teams.  e.g. ["BOS", "NYY"]
filter_teams = None

# Produce an Excel report of all changes? "yes" or "no"
produce_report = "no"
# ─────────────────────────────────────────────────────────────────────────────

# Six 2026 per-division workbooks (huronalytics account), replacing the two old
# AL/NL books. main() opens each book and walks its team tabs, so pointing at the
# six is all the routing this script needs; NLE2026 also carries ROC/AAA/FCL.
SPREADSHEET_IDS = {
    'ALE2026': '1YbgAliQzXePiFan-ruwJ50G80l4AjeyTGN8cO3KJ1XI',
    'ALC2026': '14gglESfgJoT90crQb5hHoEZNUFDZ5chPLbUIV9mlm4E',
    'ALW2026': '1eSFfKRo5kSImjP0SZ1SMssGrOhrKSZM9GOHiwntIlhs',
    'NLE2026': '1BypxxlWgQAltETOLqccOYigeo8nXX-FIuVv6rhT4anA',
    'NLC2026': '1-I8BVEw9bR9rzGVYJao_Ar0bjYZF54pi5pm3YEluB9w',
    'NLW2026': '1vm257A676FORcSRzXcNj6txgehGhYI7k5mnmsgQCYH0',
}

# Spreadsheet column name -> Statcast CSV column name
SUPPLEMENT_MAP = {
    'ArmAngle': 'arm_angle',
    'BatSpeed': 'bat_speed',
    'SwingLength': 'swing_length',
    'AttackAngle': 'attack_angle',
    'AttackDirection': 'attack_direction',
    'SwingPathTilt': 'swing_path_tilt',
    'RunExp': 'delta_pitcher_run_exp',
    'xBA': 'estimated_ba_using_speedangle',
    'xSLG': 'estimated_slg_using_speedangle',
    'xwOBA': 'estimated_woba_using_speedangle',
    'Outs': 'outs_when_up',
    'Event': 'events',
    'Barrel': 'launch_speed_angle',
}

# Swing-tracking cluster: if BatSpeed is missing or sub-50, the entire
# cluster is treated as invalid and dropped together (matches Pitcher2026).
SWING_CLUSTER_COLS = {'BatSpeed', 'SwingLength', 'AttackAngle',
                      'AttackDirection', 'SwingPathTilt'}

# Columns that store raw integer values from Statcast (no rounding needed)
INT_COLS = {'Outs', 'Barrel'}  # Raw integer values (Barrel = launch_speed_angle 1-6)

# Columns that store free-form strings (no numeric coercion, custom translator).
STRING_COLS = {'Event'}

# Columns where official Statcast data should always overwrite estimates
# (even if the cell already has a value from the initial download). Barrel is
# included because Pitcher2026 seeds it with the code_barrel estimate (6 or
# blank), which the official launch_speed_angle (1-6) should replace.
ALWAYS_OVERWRITE_COLS = {'ArmAngle', 'Barrel'}

# Columns that only ever OVERWRITE existing values; they are never used to
# fill a blank cell. Intended for scoring-change corrections (e.g., official
# scorer flips a play from hit to error), where the initial download already
# populated the cell via the MLB Stats API feed.
OVERWRITE_ONLY_COLS = {'Event'}

# Statcast `events` code -> MLB Stats API event string (the format Wally's
# sheet already stores, produced by Pitcher2026.py via play.result.event).
# Only scoring-change-relevant codes are mapped. Statcast's generic
# `field_out` is intentionally OMITTED: MLB Stats API keeps Groundout /
# Flyout / Lineout / Pop Out as distinct events, and we have no way to
# disambiguate from Statcast alone. A missing mapping means "skip; do not
# overwrite the existing sheet value."
STATCAST_TO_MLB_EVENT = {
    'single': 'Single',
    'double': 'Double',
    'triple': 'Triple',
    'home_run': 'Home Run',
    'strikeout': 'Strikeout',
    'strikeout_double_play': 'Strikeout Double Play',
    'walk': 'Walk',
    'intent_walk': 'Intent Walk',
    'hit_by_pitch': 'Hit By Pitch',
    'sac_fly': 'Sac Fly',
    'sac_fly_double_play': 'Sac Fly Double Play',
    'sac_bunt': 'Sac Bunt',
    'sac_bunt_double_play': 'Sac Bunt Double Play',
    'catcher_interf': 'Catcher Interference',
    'field_error': 'Field Error',
    'fielders_choice': 'Fielders Choice',
    'fielders_choice_out': 'Fielders Choice Out',
    'grounded_into_double_play': 'Grounded Into DP',
    'double_play': 'Double Play',
    'triple_play': 'Triple Play',
    'force_out': 'Forceout',
}

# Per-column rounding (default is 1 decimal for anything not listed)
ROUND_DECIMALS = {
    'ArmAngle': 1,
    'BatSpeed': 1,
    'SwingLength': 1,
    'AttackAngle': 1,
    'AttackDirection': 1,
    'SwingPathTilt': 1,
    'RunExp': 3,
    'xBA': 3,
    'xSLG': 3,
    'xwOBA': 3,
}

# Team abbreviation mapping: spreadsheet tab name -> Statcast Search abbreviation
STATCAST_TEAM_MAP = {
    'ATH': 'OAK',
    'KCR': 'KC',
    'SDP': 'SD',
    'SFG': 'SF',
    'TBR': 'TB',
}

MLB_TEAMS = {
    'ARI', 'ATH', 'ATL', 'BAL', 'BOS', 'CHC', 'CIN', 'CLE', 'COL', 'CWS',
    'DET', 'HOU', 'KCR', 'LAA', 'LAD', 'MIA', 'MIL', 'MIN', 'NYM', 'NYY',
    'PHI', 'PIT', 'SDP', 'SEA', 'SFG', 'STL', 'TBR', 'TEX', 'TOR', 'WSH',
}
ROC_TEAMS = {'ROC'}  # MiLB/affiliate teams tracked alongside MLB
ALL_TRACKED_TEAMS = MLB_TEAMS | ROC_TEAMS


def date_in_range(date_str):
    """Check if a date string falls within the configured range (inclusive)."""
    if start_date is None and end_date is None:
        return True
    if start_date and date_str < start_date:
        return False
    if end_date and date_str > end_date:
        return False
    return True


def download_statcast(team_tab, date_min, date_max, session):
    """Download Statcast Search CSV for a team and date range.
    Returns a dict keyed by (game_pk, at_bat_number, pitch_number) -> row dict."""
    statcast_team = STATCAST_TEAM_MAP.get(team_tab, team_tab)
    print(f"    Downloading Statcast for {team_tab} ({statcast_team}) "
          f"{date_min} to {date_max}...")

    url = "https://baseballsavant.mlb.com/statcast_search/csv"
    params = {
        'all': 'true',
        'type': 'details',
        'game_date_gt': date_min,
        'game_date_lt': date_max,
        'team': statcast_team,
        'player_type': 'pitcher',
        'min_pitches': '0',
        'min_results': '0',
        'sort_col': 'pitches',
        'sort_order': 'desc',
    }

    # Retry transient Savant errors (timeouts, 5xx, connection resets)
    # so a single hiccup doesn't drop a team's worth of supplement data.
    response = None
    for attempt in range(4):
        try:
            response = session.get(url, params=params, timeout=120)
            if response.status_code == 200:
                break
            if response.status_code >= 500 and attempt < 3:
                wait = 5 * (2 ** attempt)  # 5, 10, 20 s
                print(f"    Savant returned {response.status_code}, retrying in {wait}s...")
                time.sleep(wait)
                continue
            print(f"    Statcast returned status {response.status_code}")
            return None
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError) as e:
            if attempt < 3:
                wait = 5 * (2 ** attempt)
                print(f"    Savant request failed ({type(e).__name__}), retrying in {wait}s...")
                time.sleep(wait)
                continue
            print(f"    Timeout downloading Statcast data")
            return None

    if response is None or response.status_code != 200:
        return None

    try:
        csv_text = response.text
        if not csv_text or csv_text.strip() == '' or 'No Results' in csv_text[:100]:
            print(f"    No Statcast data available yet")
            return None

        df = pd.read_csv(StringIO(csv_text))
        if df.empty:
            print(f"    Empty DataFrame")
            return None

        # Verify merge keys exist
        for col in ['game_pk', 'at_bat_number', 'pitch_number']:
            if col not in df.columns:
                print(f"    Missing merge key: {col}")
                return None

        # Build lookup dict keyed by PitchID components
        # Use vectorised pandas ops where possible, then convert to dict
        statcast_cols = list(SUPPLEMENT_MAP.values())
        available = [c for c in statcast_cols if c in df.columns]

        # Build string keys once (vectorised)
        keys_df = pd.DataFrame({
            'k0': df['game_pk'].astype(int).astype(str),
            'k1': df['at_bat_number'].astype(int).astype(str),
            'k2': df['pitch_number'].astype(int).astype(str),
        })

        # Swing-cluster invalidity mask: BatSpeed missing or <50 means the
        # entire swing-tracking frame is unreliable; null all members together.
        if 'bat_speed' in df.columns:
            bs_numeric = pd.to_numeric(df['bat_speed'], errors='coerce')
            swing_invalid = bs_numeric.isna() | (bs_numeric < 50)
        else:
            swing_invalid = pd.Series(False, index=df.index)

        # Pre-format each supplement column into a string Series
        formatted = {}
        for sheet_col, csv_col in SUPPLEMENT_MAP.items():
            if csv_col not in df.columns:
                continue
            series = df[csv_col]
            if sheet_col in STRING_COLS:
                # String column: custom translator. For Event, translate Statcast
                # lowercase_underscore codes to MLB Stats API title-case strings
                # that Wally's sheet uses. Unmapped codes (including `field_out`)
                # are dropped so the downstream overwrite step leaves the cell
                # alone.
                if sheet_col == 'Event':
                    mapped = series.map(STATCAST_TO_MLB_EVENT)
                    formatted[sheet_col] = mapped.dropna()
                else:
                    formatted[sheet_col] = series.dropna().astype(str)
            elif sheet_col in INT_COLS:
                # Integer columns: raw int as string, NaN -> None
                s = series.dropna().astype(float).astype(int).astype(str)
                formatted[sheet_col] = s
            else:
                decimals = ROUND_DECIMALS.get(sheet_col, 1)
                numeric = pd.to_numeric(series, errors='coerce')
                if sheet_col in SWING_CLUSTER_COLS:
                    numeric = numeric.where(~swing_invalid)
                rounded = numeric.round(decimals)
                # Format to fixed decimal string; NaN rows excluded via dropna
                fmt_func = (lambda d: lambda v: f"{v:.{d}f}")(decimals)
                s = rounded.dropna().map(fmt_func)
                formatted[sheet_col] = s

        # Runners column (vectorised)
        has_runners = all(c in df.columns for c in ['on_1b', 'on_2b', 'on_3b'])
        if has_runners:
            r1 = df['on_1b'].notna()
            r2 = df['on_2b'].notna()
            r3 = df['on_3b'].notna()
            # Build runners string per row
            runners = pd.Series('0', index=df.index)
            # Assign combinations (most common first for speed)
            mask_any = r1 | r2 | r3
            if mask_any.any():
                parts = []
                for mask, label in [(r1, '1'), (r2, '2'), (r3, '3')]:
                    parts.append(mask.map({True: label, False: ''}))
                runners = (parts[0] + '+' + parts[1] + '+' + parts[2]).str.strip('+').str.replace(r'\++', '+', regex=True)
                runners = runners.replace('', '0')

        # Assemble lookup dict
        lookup = {}
        for i in df.index:
            key = (keys_df.at[i, 'k0'], keys_df.at[i, 'k1'], keys_df.at[i, 'k2'])
            data = {}
            for sheet_col in formatted:
                if i in formatted[sheet_col].index:
                    data[sheet_col] = formatted[sheet_col].at[i]
                elif sheet_col in ALWAYS_OVERWRITE_COLS:
                    data[sheet_col] = ''
            if has_runners:
                data['Runners'] = runners.at[i]
            if data:
                lookup[key] = data

        print(f"    Got {len(lookup)} pitches with supplement data "
              f"(columns: {available})")
        return lookup

    except requests.exceptions.Timeout:
        print(f"    Timeout downloading Statcast data")
        return None
    except Exception as e:
        print(f"    Error: {e}")
        return None


_TRANSIENT_SHEETS_CODES = ('429', '500', '502', '503', '504')


def _retry_sheets_call(fn, label, max_retries=5):
    """Retry a Sheets API call on rate-limit (429), transient 5xx errors, and
    network-level drops (connection reset, timeout)."""
    for attempt in range(max_retries):
        try:
            return fn()
        except gspread.exceptions.APIError as e:
            msg = str(e)
            code = next((c for c in _TRANSIENT_SHEETS_CODES if c in msg), None)
            if code and attempt < max_retries - 1:
                wait = min(60, 5 * (2 ** attempt))  # 5, 10, 20, 40, 60 s
                kind = 'Rate limited' if code == '429' else f'Transient {code}'
                print(f"    {kind} during {label}, waiting {wait}s before retry "
                      f"({attempt + 1}/{max_retries - 1})...")
                time.sleep(wait)
            else:
                raise
        except (requests.exceptions.ConnectionError,
                requests.exceptions.Timeout,
                requests.exceptions.ChunkedEncodingError) as e:
            # Transient network blip (e.g. ConnectionResetError [Errno 54] from a
            # dropped TLS connection during a write). gspread surfaces these as
            # requests errors, not APIError, so they need their own retry path.
            # The Sheets writes are idempotent (same cells and values), so
            # resending after a reset is safe even if Google applied the lost
            # request.
            if attempt < max_retries - 1:
                wait = min(60, 5 * (2 ** attempt))  # 5, 10, 20, 40, 60 s
                print(f"    Connection error during {label} "
                      f"({type(e).__name__}), waiting {wait}s before retry "
                      f"({attempt + 1}/{max_retries - 1})...")
                time.sleep(wait)
            else:
                raise


def read_sheet_with_retry(ws, max_retries=5):
    return _retry_sheets_call(ws.get_all_values, 'sheet read', max_retries)


def update_cells_with_retry(ws, cells, max_retries=5, **kwargs):
    return _retry_sheets_call(
        lambda: ws.update_cells(cells, **kwargs), 'cell write', max_retries,
    )


def write_report(report_data, output_dir='/Users/wallyhuron/Downloads/'):
    """Write an Excel report with one tab per team showing all changed rows.
    Changed cells are bold; column header row shows which columns had changes."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    bold = Font(bold=True)
    teams_with_data = sorted(t for t, rows in report_data.items() if rows)

    if not teams_with_data:
        print("  No changes to report.")
        return None

    for team in teams_with_data:
        ws = wb.create_sheet(title=team)
        entries = report_data[team]
        header = entries[0]['header']

        # Write header row (bold)
        for c, col_name in enumerate(header, start=1):
            cell = ws.cell(row=1, column=c, value=col_name)
            cell.font = bold

        # Write data rows
        for r, entry in enumerate(entries, start=2):
            row_vals = entry['row_values']
            changes = entry['changes']  # {col_idx: 'new'|'overwrite'}
            for c, val in enumerate(row_vals):
                cell = ws.cell(row=r, column=c + 1, value=val)
                if c in changes:
                    cell.font = bold

    stamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    path = os.path.join(output_dir, f'backfill_report_{stamp}.xlsx')
    wb.save(path)
    return path


def main():
    print(f"Date range: {start_date or '(all)'} to {end_date or '(all)'}")
    if filter_teams:
        print(f"Teams: {', '.join(sorted(filter_teams))}")

    # Default gspread service account (~/.config/gspread/service_account.json =
    # huronalytics), the writer on all six division books and the same account
    # the append path (sheets_append.py) uses. The old repo-local
    # service_account.json was the st-leaderboard reader, which has neither
    # access to nor write permission on the new books.
    gc = gspread.service_account()

    session = requests.Session()
    session.headers.update({
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7)',
        'Accept': 'text/csv',
    })

    total_filled = 0
    total_overwritten = 0
    report_data = {}  # team -> list of {header, row_values, changes}

    for sheet_label, sheet_id in SPREADSHEET_IDS.items():
        sh = gc.open_by_key(sheet_id)
        print(f"\n{'='*60}")
        print(f"[{sheet_label}] {sh.title}")
        print(f"{'='*60}")

        for i, ws in enumerate(sh.worksheets()):
            tab_name = ws.title.upper()

            # Skip WBC and non-tracked tabs
            if tab_name not in ALL_TRACKED_TEAMS:
                continue
            if filter_teams and tab_name not in filter_teams:
                continue

            print(f"\n[{ws.title}]")
            if i > 0:
                time.sleep(1.5)

            rows = read_sheet_with_retry(ws)
            if not rows or len(rows) < 2:
                print(f"  Empty sheet")
                continue

            header = rows[0]
            col_idx = {name: j for j, name in enumerate(header) if name}

            # Verify required columns exist
            pitch_id_col = col_idx.get('PitchID')
            if pitch_id_col is None:
                print(f"  No PitchID column — skipping")
                continue

            # Find supplement column indices (SUPPLEMENT_MAP + Runners)
            supp_col_idx = {}
            for sheet_col in SUPPLEMENT_MAP:
                if sheet_col in col_idx:
                    supp_col_idx[sheet_col] = col_idx[sheet_col]
            if 'Runners' in col_idx:
                supp_col_idx['Runners'] = col_idx['Runners']
            if not supp_col_idx:
                print(f"  No supplement columns found — skipping")
                continue

            # Find rows that need filling:
            # PitchID exists AND (at least one supplement column is empty
            # OR has an always-overwrite column that might contain an estimate)
            needs_fill = []  # (row_index_1based, pitch_id, cols_to_update)
            # cols_to_update entries: (sheet_col, existing_value) — empty string means new fill
            game_dates = set()
            date_col = col_idx.get('Game Date')

            for r_idx, row in enumerate(rows[1:], start=2):
                pid = row[pitch_id_col] if pitch_id_col < len(row) else ''
                if not pid or '_' not in pid:
                    continue

                # Apply date range filter for supplement backfill
                if date_col is not None:
                    gd = row[date_col] if date_col < len(row) else ''
                    if not date_in_range(gd):
                        continue

                # Check which supplement columns need updating:
                # - Empty columns need filling (but NOT for OVERWRITE_ONLY_COLS)
                # - ALWAYS_OVERWRITE_COLS need updating even if they have a value
                #   (the existing value may be an estimate that official data should replace)
                # - OVERWRITE_ONLY_COLS update existing values only (for scoring
                #   corrections like hit↔error); never used to fill a blank cell.
                cols_to_update = []
                for sheet_col, c_idx in supp_col_idx.items():
                    val = row[c_idx] if c_idx < len(row) else ''
                    is_empty = (val == '' or val is None)
                    if is_empty and sheet_col in OVERWRITE_ONLY_COLS:
                        continue
                    if is_empty:
                        cols_to_update.append((sheet_col, ''))
                    elif sheet_col in ALWAYS_OVERWRITE_COLS or sheet_col in OVERWRITE_ONLY_COLS:
                        cols_to_update.append((sheet_col, val))

                if cols_to_update:
                    needs_fill.append((r_idx, pid, cols_to_update))
                    if date_col is not None:
                        gd = row[date_col] if date_col < len(row) else ''
                        if gd:
                            game_dates.add(gd)

            if not needs_fill:
                print(f"  All rows filled — nothing to do")
                continue

            print(f"  {len(needs_fill)} rows need supplement data "
                  f"(dates: {sorted(game_dates)})")

            # Download Statcast data for this team's date range
            if not game_dates:
                print(f"  No game dates found — skipping")
                continue

            date_min = min(game_dates)
            date_max = max(game_dates)

            time.sleep(3)  # Be polite to Baseball Savant
            lookup = download_statcast(ws.title, date_min, date_max, session)

            if lookup is None:
                print(f"  No Statcast data available — skipping")
                continue

            # Match and prepare cell updates
            cells_to_update = []
            new_fill_cells = 0
            overwrite_cells = 0
            team_report_rows = []

            for r_idx, pid, cols_to_update in needs_fill:
                # Split PitchID: game_pk_atbat(zero-padded)_pitch(zero-padded)
                # Strip padding to match Statcast lookup keys (unpadded ints)
                parts = pid.split('_')
                if len(parts) != 3:
                    continue
                key = (parts[0], str(int(parts[1])), str(int(parts[2])))

                statcast_row = lookup.get(key, {})
                row_changes = {}  # col_idx -> 'new'|'overwrite'

                for sheet_col, existing_val in cols_to_update:
                    if sheet_col in statcast_row:
                        val = statcast_row[sheet_col]
                        # Don't write empty values for overwrite cols —
                        # that would erase data when official data isn't ready yet
                        if not val and (sheet_col in ALWAYS_OVERWRITE_COLS
                                        or sheet_col in OVERWRITE_ONLY_COLS):
                            continue
                        # Skip if overwrite value is identical to existing
                        if existing_val and str(val) == str(existing_val):
                            continue
                        cell = gspread.Cell(
                            row=r_idx,
                            col=supp_col_idx[sheet_col] + 1,  # 1-indexed
                            value=val,
                        )
                        cells_to_update.append(cell)
                        c_idx = col_idx[sheet_col]
                        if existing_val:
                            overwrite_cells += 1
                            row_changes[c_idx] = 'overwrite'
                        else:
                            new_fill_cells += 1
                            row_changes[c_idx] = 'new'

                # Collect report data for this row if anything changed
                if row_changes and produce_report == 'yes':
                    # Build row with new values applied
                    row_vals = list(rows[r_idx - 1])
                    for c_idx, change_type in row_changes.items():
                        # Find the cell we're about to write for this column
                        for cell in cells_to_update:
                            if cell.row == r_idx and cell.col == c_idx + 1:
                                row_vals[c_idx] = cell.value
                                break
                    team_report_rows.append({
                        'header': header,
                        'row_values': row_vals,
                        'changes': row_changes,
                    })

            if cells_to_update:
                print(f"  Writing {len(cells_to_update)} cells "
                      f"({new_fill_cells} new, {overwrite_cells} overwritten)...")
                update_cells_with_retry(ws, cells_to_update, value_input_option='RAW')
                total_filled += new_fill_cells
                total_overwritten += overwrite_cells
                time.sleep(2)  # Rate limit buffer after write
            else:
                print(f"  No new data to fill, no overwrites changed.")

            if team_report_rows:
                report_data[tab_name] = team_report_rows

    parts = []
    if total_filled:
        parts.append(f"{total_filled} new cells filled")
    if total_overwritten:
        parts.append(f"{total_overwritten} cells overwritten")
    if parts:
        print(f"\nDone. {', '.join(parts)}.")
    else:
        print(f"\nDone. No new data added, no data overwritten.")

    if produce_report == 'yes':
        report_path = write_report(report_data)
        if report_path:
            print(f"Report saved to: {report_path}")


if __name__ == '__main__':
    # ── CLI overrides (optional — edit start_date/end_date at top of file as before) ──
    parser = argparse.ArgumentParser(description='Backfill supplemental Statcast data into Google Sheet')
    parser.add_argument('--start', default=None, help='Start date YYYY-MM-DD, or "none" for all dates')
    parser.add_argument('--end', default=None, help='End date YYYY-MM-DD, or "none" for all dates')
    parser.add_argument('--teams', default=None, help='Comma-separated team abbreviations (e.g., BOS,NYY)')
    parser.add_argument('--report', default=None, help='"yes" to produce an Excel report of changes')
    args = parser.parse_args()

    # Only override module-level globals if CLI args were explicitly passed
    if args.start is not None:
        start_date = None if args.start.lower() == 'none' else args.start
    if args.end is not None:
        end_date = None if args.end.lower() == 'none' else args.end
    if args.teams is not None:
        filter_teams = [t.strip().upper() for t in args.teams.split(',') if t.strip()]
    if args.report is not None:
        produce_report = args.report.lower()

    main()
